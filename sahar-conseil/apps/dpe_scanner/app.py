"""
SAHAR Conseil — DPE Scanner v2
Détection et prospection des passoires thermiques F/G.
Données ADEME. CRM intégré. Mobile-first.
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, date

from shared.auth import verifier_acces
from shared import crm_db

st.set_page_config(
    page_title="DPE Scanner — SAHAR Conseil",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    .main .block-container{padding-top:1rem;padding-bottom:1rem}
    .stMetric label{font-size:.78rem;color:#73726c}
    .stMetric [data-testid="stMetricValue"]{font-size:1.4rem}
    div[data-testid="stSidebarContent"]{padding-top:.75rem}
    .stTabs [data-baseweb="tab"]{font-size:.85rem}
</style>
""", unsafe_allow_html=True)

verifier_acces()
crm_db.init_crm()

# ─── CHARGEMENT DPE ──────────────────────────────────────────────────────────

DPE_API = "https://data.ademe.fr/data-fair/api/v1/datasets/dpe-v2-logements-existants/lines"

@st.cache_data(ttl=3600, show_spinner="Interrogation ADEME...")
def fetch_dpe(code_postal: str, nb: int = 1000) -> pd.DataFrame:
    import requests

    # URLs ADEME — dataset renommé "dpe03existant" en 2024
    ADEME_URLS = [
        "https://data.ademe.fr/data-fair/api/v1/datasets/dpe03existant/lines",
        "https://data.ademe.fr/data-fair/api/v1/datasets/dpe-v2-logements-existants/lines",
        "https://data.ademe.fr/data-fair/api/v1/datasets/dpe-v2-logements-existants-2/lines",
    ]

    SELECT = (
        "numero_dpe,date_etablissement_dpe,etiquette_dpe,etiquette_ges,"
        "conso_5_usages_e_finale,emission_ges_5_usages,"
        "adresse_ban,code_postal_ban,nom_commune_ban,"
        "latitude,longitude,type_batiment,annee_construction,"
        "surface_habitable_logement,type_energie_principale_chauffage"
    )

    params = {
        "q": code_postal,
        "q_fields": "code_postal_ban",
        "size": nb,
        "select": SELECT,
    }

    for url in ADEME_URLS:
        try:
            r = requests.get(url, params=params, timeout=30)
            if r.status_code == 200:
                data = r.json().get("results", [])
                if data:
                    df = pd.DataFrame(data)
                    df["date_etablissement_dpe"] = pd.to_datetime(
                        df.get("date_etablissement_dpe", pd.Series(dtype=str)), errors="coerce"
                    )
                    for col in ["conso_5_usages_e_finale", "emission_ges_5_usages",
                                "surface_habitable_logement", "latitude", "longitude"]:
                        if col in df.columns:
                            df[col] = pd.to_numeric(df[col], errors="coerce")
                    if "annee_construction" in df.columns:
                        df["annee_construction"] = pd.to_numeric(df["annee_construction"], errors="coerce").astype("Int64")
                    return df
        except Exception as e:
            continue

    st.error(f"API ADEME indisponible — réessayez dans quelques minutes.")
    return pd.DataFrame()


def score_urgence(df: pd.DataFrame) -> pd.Series:
    """Score 0–100 : étiquette DPE (50%) + conso (30%) + ancienneté (20%)."""
    mapping = {"G":100,"F":75,"E":50,"D":30,"C":15,"B":5,"A":0}
    s_etiq = df["etiquette_dpe"].map(mapping).fillna(30)

    conso = df["conso_5_usages_e_finale"].fillna(300).clip(0, 800)
    s_conso = (conso / 800 * 100)

    annee = df["annee_construction"].fillna(1980)
    s_anc = ((2024 - annee) / 100 * 100).clip(0, 100)

    return (s_etiq * 0.5 + s_conso * 0.3 + s_anc * 0.2).round(0).clip(0, 100).astype(int)


# ─── SIDEBAR ─────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("### ⚡ DPE Scanner")
    st.caption("SAHAR Conseil")
    st.markdown("---")

    # Multi codes postaux
    cp_input = st.text_area(
        "Codes postaux (un par ligne)",
        value="33000\n33100\n33200",
        height=100,
        help="Entrer un ou plusieurs codes postaux"
    )
    codes_postaux = [c.strip() for c in cp_input.strip().splitlines() if c.strip().isdigit() and len(c.strip()) == 5]

    etiquettes = st.multiselect(
        "Étiquettes cibles",
        ["G","F","E","D","C","B","A"],
        default=["G","F"],
    )

    types_bâtiment = st.multiselect(
        "Type de bâtiment",
        ["Maison individuelle","Appartement","Immeuble","Local commercial","Autre"],
        default=["Maison individuelle","Appartement"],
    )

    score_min = st.slider("Score urgence minimum", 0, 100, 40, 5)
    nb_par_cp = st.slider("Résultats par code postal", 100, 2000, 500, 100)

    st.markdown("---")
    if st.button("🔄 Vider le cache"):
        st.cache_data.clear()
        st.rerun()
    if st.button("🔓 Déconnexion"):
        st.session_state.pop("auth_ok", None)
        st.rerun()

# ─── CHARGEMENT ──────────────────────────────────────────────────────────────

st.title("⚡ DPE Scanner")
st.caption("Détection de passoires thermiques — Source ADEME")

if not codes_postaux:
    st.warning("Entrez au moins un code postal valide.")
    st.stop()

dfs = []
progress = st.progress(0, text="Chargement...")
for i, cp in enumerate(codes_postaux):
    progress.progress((i+1)/len(codes_postaux), text=f"Chargement {cp}...")
    df_cp = fetch_dpe(cp, nb=nb_par_cp)
    if not df_cp.empty:
        dfs.append(df_cp)
progress.empty()

if not dfs:
    st.warning("Aucun résultat. Vérifiez les codes postaux.")
    st.stop()

df = pd.concat(dfs, ignore_index=True).drop_duplicates(subset=["numero_dpe"])

# Filtres
if etiquettes:
    df = df[df["etiquette_dpe"].isin(etiquettes)]
if types_bâtiment and "type_batiment" in df.columns:
    df = df[df["type_batiment"].isin(types_bâtiment)]

if df.empty:
    st.info("Aucun résultat avec ces filtres.")
    st.stop()

df["score"] = score_urgence(df)
df = df[df["score"] >= score_min].sort_values("score", ascending=False)

# ─── KPIs ────────────────────────────────────────────────────────────────────

k1,k2,k3,k4,k5 = st.columns(5)
with k1: st.metric("Logements", f"{len(df):,}")
with k2: st.metric("Classe G", int((df["etiquette_dpe"]=="G").sum()))
with k3: st.metric("Classe F", int((df["etiquette_dpe"]=="F").sum()))
with k4: st.metric("Score moyen", f"{df['score'].mean():.0f}/100")
with k5:
    conso_med = df["conso_5_usages_e_finale"].median()
    st.metric("Conso médiane", f"{conso_med:.0f} kWh/m²" if not np.isnan(conso_med) else "—")

st.markdown("---")

# ─── ONGLETS ─────────────────────────────────────────────────────────────────

STAGES = ["Détecté","Contacté","Qualifié","Proposition","Closing"]
STAGE_COLORS = {"Détecté":"#e5e5e5","Contacté":"#fff3cd","Qualifié":"#cfe2ff",
                "Proposition":"#d1ecf1","Closing":"#d4edda"}

tab_liste, tab_carte, tab_stats, tab_crm, tab_export = st.tabs([
    "📋 Prospects", "🗺️ Carte", "📊 Stats", "💼 CRM", "📥 Export"
])

# ── LISTE ──────────────────────────────────────────────────────────────────

with tab_liste:
    col_left, col_right = st.columns([3,1])

    with col_right:
        commune_opts = ["Toutes"] + sorted(df["nom_commune_ban"].dropna().unique().tolist())
        commune_sel = st.selectbox("Commune", commune_opts)
        tri = st.selectbox("Trier par", ["score","conso_5_usages_e_finale","annee_construction","surface_habitable_logement"])
        nb_aff = st.slider("Nb lignes", 10, 500, 50, 10)

    with col_left:
        df_aff = df.copy()
        if commune_sel != "Toutes":
            df_aff = df_aff[df_aff["nom_commune_ban"] == commune_sel]
        df_aff = df_aff.sort_values(tri, ascending=(tri != "score")).head(nb_aff)

        cols_show = [c for c in [
            "score","etiquette_dpe","etiquette_ges","adresse_ban","nom_commune_ban",
            "code_postal_ban","type_batiment","annee_construction",
            "surface_habitable_logement","conso_5_usages_e_finale",
            "type_energie_principale_chauffage","date_etablissement_dpe","numero_dpe"
        ] if c in df_aff.columns]

        df_table = df_aff[cols_show].copy()
        if "date_etablissement_dpe" in df_table.columns:
            df_table["date_etablissement_dpe"] = df_table["date_etablissement_dpe"].dt.strftime("%d/%m/%Y")

        df_table = df_table.rename(columns={
            "score":"Score","etiquette_dpe":"DPE","etiquette_ges":"GES",
            "adresse_ban":"Adresse","nom_commune_ban":"Commune","code_postal_ban":"CP",
            "type_batiment":"Type","annee_construction":"Année",
            "surface_habitable_logement":"Surface m²",
            "conso_5_usages_e_finale":"Conso kWh/m²",
            "type_energie_principale_chauffage":"Énergie",
            "date_etablissement_dpe":"Date DPE","numero_dpe":"N° DPE"
        })

        st.dataframe(
            df_table, use_container_width=True, hide_index=True,
            column_config={
                "Score": st.column_config.ProgressColumn("Score", min_value=0, max_value=100, format="%d"),
                "DPE": st.column_config.TextColumn("DPE", width="small"),
                "Surface m²": st.column_config.NumberColumn("m²", format="%.0f"),
                "Conso kWh/m²": st.column_config.NumberColumn("kWh/m²", format="%.0f"),
            }
        )
        st.caption(f"{len(df_aff):,} prospects • {len(df):,} total")

        # Ajout CRM rapide
        with st.expander("➕ Ajouter au CRM"):
            if len(df_aff) > 0:
                idx = st.selectbox("Sélectionner",
                    range(min(nb_aff, len(df_aff))),
                    format_func=lambda i: f"{df_aff.iloc[i]['adresse_ban']} — {df_aff.iloc[i]['etiquette_dpe']} — Score {df_aff.iloc[i]['score']}" if i < len(df_aff) else "")
                row = df_aff.iloc[idx]

                st.markdown(f"**{row['adresse_ban']}, {row['nom_commune_ban']} {row['code_postal_ban']}**")
                st.markdown(f"DPE {row['etiquette_dpe']} — {row.get('surface_habitable_logement',0):.0f} m² — Conso {row.get('conso_5_usages_e_finale',0):.0f} kWh/m² — Score **{row['score']}/100**")

                contact_opts = ["Nouveau contact"] + [f"{c['id']} — {c['nom']}" for c in st.session_state.crm_contacts]
                contact_choix = st.selectbox("Contact", contact_opts)

                if contact_choix == "Nouveau contact":
                    nom_c = st.text_input("Nom du propriétaire")
                    tel_c = st.text_input("Téléphone")
                    email_c = st.text_input("Email")
                    type_c = st.selectbox("Type", ["Propriétaire","Bailleur","Copropriété","Autre"])

                if st.button("Ajouter au CRM", type="primary"):
                    if contact_choix == "Nouveau contact" and nom_c:
                        crm_db.add_contact(nom_c, email_c, tel_c, type_c)
                        contact_id = st.session_state.crm_contacts[-1]["id"]
                    elif contact_choix != "Nouveau contact":
                        contact_id = contact_choix.split(" — ")[0]
                    else:
                        st.warning("Entrez un nom de contact.")
                        st.stop()

                    crm_db.add_opportunite(
                        contact_id=contact_id,
                        titre=f"DPE {row['etiquette_dpe']} — {row['adresse_ban']}, {row['nom_commune_ban']}",
                        adresse=f"{row['adresse_ban']}, {row['nom_commune_ban']} {row['code_postal_ban']}",
                        type_bien=row.get("type_batiment","Logement"),
                        surface=float(row.get("surface_habitable_logement",0) or 0),
                        prix=0.0,
                        prix_m2=0.0,
                        score=int(row["score"]),
                        source="DPE"
                    )
                    st.success("✓ Ajouté au CRM !")
                    st.rerun()

# ── CARTE ──────────────────────────────────────────────────────────────────

with tab_carte:
    df_geo = df.dropna(subset=["lat","lon"])
    if df_geo.empty:
        st.info("Coordonnées GPS non disponibles pour ce code postal.")
    else:
        try:
            import folium
            from streamlit_folium import st_folium

            max_pts = st.slider("Points sur la carte", 100, 2000, 500, 100)
            df_map = df_geo.head(max_pts)

            m = folium.Map(
                location=[df_map["lat"].median(), df_map["lon"].median()],
                zoom_start=13, tiles="CartoDB positron"
            )

            couleurs_dpe = {"G":"#E24B4A","F":"#BA7517","E":"#EF9F27",
                           "D":"#1D9E75","C":"#185FA5","B":"#0C447C","A":"#042C53"}

            for _, r in df_map.iterrows():
                col = couleurs_dpe.get(str(r.get("etiquette_dpe","")),"#888")
                popup = f"""<div style='font-family:sans-serif;font-size:12px;min-width:200px'>
                <b>{r.get('adresse_ban','—')}</b><br>
                {r.get('nom_commune_ban','—')} {r.get('code_postal_ban','')}<br>
                <b>DPE :</b> <span style='color:{col};font-weight:700'>{r.get('etiquette_dpe','—')}</span>
                — GES {r.get('etiquette_ges','—')}<br>
                <b>Type :</b> {r.get('type_batiment','—')}<br>
                <b>Surface :</b> {r.get('surface_habitable_logement',0):.0f} m²<br>
                <b>Construction :</b> {int(r.get('annee_construction',0)) if r.get('annee_construction') else '—'}<br>
                <b>Conso :</b> {r.get('conso_5_usages_e_finale',0):.0f} kWh/m²/an<br>
                <b>Énergie :</b> {r.get('type_energie_principale_chauffage','—')}<br>
                <b style='color:{col}'>Score urgence : {r.get('score',0)}/100</b>
                </div>"""
                folium.CircleMarker(
                    location=[r["lat"], r["lon"]],
                    radius=6,
                    color=col, fill=True, fill_color=col, fill_opacity=0.8,
                    popup=folium.Popup(popup, max_width=270),
                    tooltip=f"DPE {r.get('etiquette_dpe','—')} — Score {r.get('score',0)}"
                ).add_to(m)

            # Légende
            cols_leg = st.columns(7)
            for i, (etiq, col) in enumerate(couleurs_dpe.items()):
                with cols_leg[i]:
                    nb = int((df["etiquette_dpe"]==etiq).sum())
                    st.markdown(f'<div style="text-align:center"><span style="color:{col};font-weight:700;font-size:1.1rem">{etiq}</span><br><span style="font-size:.75rem;color:#888">{nb}</span></div>', unsafe_allow_html=True)

            st_folium(m, height=500, use_container_width=True)

        except ImportError:
            st.warning("Installer folium et streamlit-folium.")

# ── STATS ──────────────────────────────────────────────────────────────────

with tab_stats:
    try:
        import plotly.express as px
        import plotly.graph_objects as go

        c1, c2 = st.columns(2)
        with c1:
            dist = df["etiquette_dpe"].value_counts().reset_index()
            dist.columns = ["DPE","Nb"]
            fig = px.bar(dist, x="DPE", y="Nb", title="Répartition étiquettes",
                         color="DPE",
                         color_discrete_map={"G":"#E24B4A","F":"#BA7517","E":"#EF9F27","D":"#1D9E75","C":"#185FA5","B":"#0C447C","A":"#042C53"})
            fig.update_layout(plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", showlegend=False, height=280)
            st.plotly_chart(fig, use_container_width=True)

        with c2:
            if "annee_construction" in df.columns:
                df_a = df.dropna(subset=["annee_construction"])
                fig2 = px.histogram(df_a, x="annee_construction", nbins=20,
                                    title="Année de construction",
                                    color_discrete_sequence=["#185FA5"])
                fig2.update_layout(plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", height=280)
                st.plotly_chart(fig2, use_container_width=True)

        c3, c4 = st.columns(2)
        with c3:
            if "conso_5_usages_e_finale" in df.columns:
                fig3 = px.box(df, x="etiquette_dpe", y="conso_5_usages_e_finale",
                              title="Consommation par étiquette (kWh/m²/an)",
                              color="etiquette_dpe",
                              color_discrete_map={"G":"#E24B4A","F":"#BA7517","E":"#EF9F27","D":"#1D9E75","C":"#185FA5"})
                fig3.update_layout(plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", showlegend=False, height=280)
                st.plotly_chart(fig3, use_container_width=True)

        with c4:
            top_comm = df.groupby("nom_commune_ban").size().nlargest(10).reset_index()
            top_comm.columns = ["Commune","Nb passoires"]
            fig4 = px.bar(top_comm, x="Nb passoires", y="Commune", orientation="h",
                          title="Top communes — passoires F/G",
                          color_discrete_sequence=["#E24B4A"])
            fig4.update_layout(plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", height=280)
            st.plotly_chart(fig4, use_container_width=True)

    except ImportError:
        st.info("Installer plotly.")

# ── CRM ────────────────────────────────────────────────────────────────────

with tab_crm:

    def action_buttons_dpe(contact: dict) -> None:
        tel = contact.get("tel","").strip()
        email = contact.get("email","").strip()
        nom = contact.get("nom","")
        btns = []
        if tel:
            btns.append(f'<a href="tel:{tel}" style="display:inline-flex;align-items:center;gap:.4rem;background:#1D9E75;color:#fff;padding:.55rem 1rem;border-radius:6px;text-decoration:none;font-size:.85rem;font-weight:500">📞 Appeler</a>')
            btns.append(f'<a href="sms:{tel}" style="display:inline-flex;align-items:center;gap:.4rem;background:#185FA5;color:#fff;padding:.55rem 1rem;border-radius:6px;text-decoration:none;font-size:.85rem;font-weight:500">💬 SMS</a>')
            wa = tel.replace("+","").replace(" ","")
            btns.append(f'<a href="https://wa.me/{wa}" target="_blank" style="display:inline-flex;align-items:center;gap:.4rem;background:#25D366;color:#fff;padding:.55rem 1rem;border-radius:6px;text-decoration:none;font-size:.85rem;font-weight:500">🟢 WhatsApp</a>')
        if email:
            btns.append(f'<a href="mailto:{email}?subject=Rénovation énergétique — {nom}" style="display:inline-flex;align-items:center;gap:.4rem;background:#444;color:#fff;padding:.55rem 1rem;border-radius:6px;text-decoration:none;font-size:.85rem;font-weight:500">✉️ Email</a>')
        if btns:
            st.markdown('<div style="display:flex;gap:.5rem;flex-wrap:wrap;margin:.5rem 0">' + "".join(btns) + '</div>', unsafe_allow_html=True)

    crm_t = st.tabs(["📌 Pipeline DPE","👥 Contacts","📅 Activités"])

    with crm_t[0]:
        opps_dpe = [o for o in st.session_state.crm_opportunites if o.get("source")=="DPE"]
        if not opps_dpe:
            st.info("Ajoutez des prospects depuis l'onglet Prospects.")
        else:
            for stage in STAGES:
                stage_opps = [o for o in opps_dpe if o["stage"]==stage]
                if stage_opps:
                    st.markdown(f'<div style="background:{STAGE_COLORS[stage]};padding:.6rem .9rem;border-radius:6px;margin:.4rem 0;font-size:.88rem"><b>{stage}</b> — {len(stage_opps)} prospect(s)</div>', unsafe_allow_html=True)
                    for opp in stage_opps:
                        contact = next((c for c in st.session_state.crm_contacts if c["id"]==opp.get("contact_id")), {})
                        with st.expander(f"{opp['titre'][:50]}", expanded=False):
                            st.markdown(f"**Adresse :** {opp['adresse']}")
                            st.markdown(f"**Surface :** {opp['surface']:.0f} m² | **Score urgence :** {opp['score']}/100")
                            st.markdown(f"**Contact :** {contact.get('nom','—')} {contact.get('tel','')}")
                            action_buttons_dpe(contact)
                            cur = STAGES.index(opp["stage"])
                            new_s = st.selectbox("Étape", STAGES, index=cur, key=f"ds_{opp['id']}")
                            if new_s != opp["stage"]:
                                crm_db.update_stage(opp["id"], new_s)
                                st.rerun()
                            act_t = st.selectbox("Activité", ["Appel","SMS","WhatsApp","Email","Visite","Devis","Note"], key=f"da_{opp['id']}")
                            act_n = st.text_input("Note", key=f"dn_{opp['id']}")
                            if st.button("✓ Enregistrer", key=f"db_{opp['id']}", type="primary"):
                                crm_db.add_activite(opp["id"], act_t, act_n)
                                st.success("✓ Activité enregistrée")
                                st.rerun()

        if opps_dpe:
            st.markdown("---")
            k1,k2,k3 = st.columns(3)
            with k1: st.metric("Prospects DPE", len(opps_dpe))
            with k2:
                sc_moy = sum(o["score"] for o in opps_dpe)/len(opps_dpe)
                st.metric("Score moyen", f"{sc_moy:.0f}/100")
            with k3: st.metric("En closing", len([o for o in opps_dpe if o["stage"]=="Closing"]))

    with crm_t[1]:
        with st.expander("➕ Nouveau contact"):
            with st.form("f_c_dpe"):
                n = st.text_input("Nom *")
                c1,c2 = st.columns(2)
                with c1: t = st.text_input("Téléphone")
                with c2: e = st.text_input("Email")
                tp = st.selectbox("Type",["Propriétaire","Bailleur","Copropriété","Syndic","Autre"])
                nt = st.text_area("Notes", height=60)
                if st.form_submit_button("Créer", type="primary") and n:
                    crm_db.add_contact(n,e,t,tp,nt)
                    st.success(f"✓ {n} créé")
                    st.rerun()
        for c in st.session_state.crm_contacts:
            with st.expander(f"**{c['nom']}** — {c['type']}"):
                st.markdown(f"📞 {c.get('tel','—')}  |  ✉️ {c.get('email','—')}")
                action_buttons_dpe(c)

    with crm_t[2]:
        acts = st.session_state.crm_activites
        if acts:
            for a in sorted(acts, key=lambda x: x["date_creation"], reverse=True)[:20]:
                opp_ref = next((o["titre"][:30] for o in st.session_state.crm_opportunites if o["id"]==a["opp_id"]), a["opp_id"])
                icon = {"Fait":"✅","À faire":"⏳","Annulé":"❌"}.get(a.get("statut",""),"•")
                st.markdown(f'<div style="border:1px solid #e5e5e5;border-radius:6px;padding:.7rem .9rem;margin:.3rem 0;font-size:.85rem"><b>{icon} {a["type"]}</b> · {a["date"]} · <span style="color:#888">{opp_ref}</span><div style="color:#444;margin-top:.25rem">{a.get("notes","")}</div></div>', unsafe_allow_html=True)
        else:
            st.info("Aucune activité.")

# ── EXPORT ─────────────────────────────────────────────────────────────────

with tab_export:
    st.subheader("Export prospects")
    c1, c2 = st.columns(2)
    with c1:
        st.caption(f"{len(df):,} prospects avec les filtres actuels")
        cols_exp = [c for c in ["score","etiquette_dpe","etiquette_ges","adresse_ban",
                                 "nom_commune_ban","code_postal_ban","type_batiment",
                                 "annee_construction","surface_habitable_logement",
                                 "conso_5_usages_e_finale","type_energie_principale_chauffage",
                                 "lat","lon","numero_dpe"] if c in df.columns]
        df_exp = df[cols_exp].copy()
        csv = df_exp.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
        st.download_button("📥 CSV Prospects",
            data=csv,
            file_name=f"sahar_dpe_{'_'.join(codes_postaux[:3])}_{datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv")

        try:
            import io, openpyxl
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                df_exp.rename(columns={
                    "score":"Score urgence","etiquette_dpe":"DPE","etiquette_ges":"GES",
                    "adresse_ban":"Adresse","nom_commune_ban":"Commune","code_postal_ban":"CP",
                    "type_batiment":"Type","annee_construction":"Année",
                    "surface_habitable_logement":"Surface m²",
                    "conso_5_usages_e_finale":"Conso kWh/m²",
                    "type_energie_principale_chauffage":"Énergie chauffage",
                    "numero_dpe":"N° DPE"
                }).to_excel(writer, sheet_name="Prospects DPE", index=False)
                ws = writer.sheets["Prospects DPE"]
                from openpyxl.styles import Font, PatternFill
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="E24B4A", end_color="E24B4A", fill_type="solid")
            buf.seek(0)
            st.download_button("📥 Excel Prospects",
                data=buf.getvalue(),
                file_name=f"sahar_dpe_{'_'.join(codes_postaux[:3])}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except ImportError:
            pass

    with c2:
        st.caption("Export CRM")
        opps_dpe = [o for o in st.session_state.crm_opportunites if o.get("source")=="DPE"]
        if opps_dpe:
            csv_crm = pd.DataFrame(opps_dpe).to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
            st.download_button("📥 CSV Pipeline DPE", data=csv_crm,
                file_name=f"sahar_crm_dpe_{datetime.now().strftime('%Y%m%d')}.csv", mime="text/csv")
        else:
            st.info("Aucune opportunité DPE dans le CRM.")
        st.metric("Prospects total", f"{len(df):,}")
        st.metric("Score ≥ 70", int((df["score"]>=70).sum()))
        st.metric("Classes G", int((df["etiquette_dpe"]=="G").sum()))

st.markdown("---")
st.caption("SAHAR Conseil — Source : ADEME Base DPE. Données officielles des diagnostics énergétiques.")
