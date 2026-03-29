"""
SAHAR Conseil — DVF Analyse Pro v2
Analyse marché immobilier + CRM Pipeline intégré
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

import streamlit as st
import pandas as pd
import numpy as np
import json
from datetime import datetime, date

from shared import crm_db
from shared.scoring_commune import compute_score_commune
from shared.rapport_pdf import generer_rapport_commune

# ─── CONFIG ──────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="SAHAR Conseil — DVF Analyse Pro",
    page_icon="🏠",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    .main .block-container{padding-top:1rem;padding-bottom:1rem}
    .stMetric label{font-size:.78rem;color:#73726c}
    .stMetric [data-testid="stMetricValue"]{font-size:1.4rem}
    div[data-testid="stSidebarContent"]{padding-top:.75rem}
    .stTabs [data-baseweb="tab"]{font-size:.85rem}
    [data-testid="stExpander"]{border:1px solid #e5e5e5;border-radius:8px}
</style>
""", unsafe_allow_html=True)

# ─── AUTH ─────────────────────────────────────────────────────────────────────

def check_auth():
    if st.session_state.get("auth_ok"):
        return True
    try:
        pwd_attendu = st.secrets["APP_PWD"]
    except Exception:
        return True  # dev sans secrets

    with st.sidebar:
        st.markdown("### 🔐 Accès SAHAR")
        pwd = st.text_input("Mot de passe", type="password", label_visibility="collapsed")
        if pwd:
            if pwd == pwd_attendu:
                st.session_state["auth_ok"] = True
                st.rerun()
            else:
                st.error("Mot de passe incorrect")
                st.stop()
        else:
            st.info("Entrez votre mot de passe")
            st.stop()

check_auth()

# ─── CHARGEMENT DONNÉES ───────────────────────────────────────────────────────

def find_dvf_file(dept: str) -> Path | None:
    """Cherche le fichier DVF dans tous les emplacements possibles."""
    bases = [
        Path(__file__).resolve().parents[2],
        Path(__file__).resolve().parents[3],
        Path("/mount/src/sahar-conseil/sahar-conseil"),
        Path("/mount/src/sahar-conseil"),
    ]
    for base in bases:
        # Parquet d'abord (plus rapide)
        p = base / "data" / "processed" / f"dvf_{dept}.parquet"
        if p.exists() and p.stat().st_size > 1000:
            return p
        # CSV ensuite
        c = base / "data" / "raw" / f"dvf_{dept}.csv"
        if c.exists() and c.stat().st_size > 1000:
            return c
    return None


@st.cache_data(ttl=3600, show_spinner=False)
def load_dvf(dept: str) -> pd.DataFrame:
    """Charge et nettoie les données DVF — parquet prioritaire."""
    path = find_dvf_file(dept)

    if path is None:
        st.error(f"Fichier DVF introuvable pour le département {dept}.")
        st.stop()

    if path.suffix == ".parquet":
        return pd.read_parquet(path)

    # CSV → traitement complet
    cols = ['id_mutation','date_mutation','nature_mutation','valeur_fonciere',
            'adresse_numero','adresse_nom_voie','code_postal','code_commune',
            'nom_commune','type_local','surface_reelle_bati','nombre_pieces_principales',
            'surface_terrain','lot1_surface_carrez','longitude','latitude','id_parcelle']
    dtypes = {'code_commune':'str','code_postal':'str',
              'nature_mutation':'category','type_local':'category','nom_commune':'category'}

    df = pd.read_csv(path, usecols=[c for c in cols if c in pd.read_csv(path, nrows=0).columns],
                     dtype=dtypes, low_memory=False)

    df['date_mutation'] = pd.to_datetime(df['date_mutation'], errors='coerce')
    for col in ['valeur_fonciere','surface_reelle_bati','surface_terrain',
                'lot1_surface_carrez','longitude','latitude']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    df = df[df['type_local'].isin(['Appartement','Maison'])]
    df = df[df['nature_mutation'] == 'Vente']
    df = df.dropna(subset=['valeur_fonciere','surface_reelle_bati'])
    df = df[(df['surface_reelle_bati'] > 5) & (df['valeur_fonciere'] > 1000)]

    df['surface_utile'] = df['surface_reelle_bati']
    if 'lot1_surface_carrez' in df.columns:
        m = df['lot1_surface_carrez'].notna() & (df['lot1_surface_carrez'] > 0)
        df.loc[m, 'surface_utile'] = df.loc[m, 'lot1_surface_carrez']

    df['prix_m2'] = (df['valeur_fonciere'] / df['surface_utile']).round(0)
    df = df[df['prix_m2'].between(500, 25000)]
    df['adresse'] = (df.get('adresse_numero', pd.Series('')).fillna('').astype(str).str.strip() + ' ' +
                     df.get('adresse_nom_voie', pd.Series('')).fillna('').astype(str)).str.strip()
    df['annee'] = df['date_mutation'].dt.year
    df['mois'] = df['date_mutation'].dt.to_period('M').astype(str)

    return df.reset_index(drop=True)


def score_opportunite(df: pd.DataFrame, w_prix=0.4, w_vol=0.3, w_dyn=0.3) -> pd.Series:
    def norm(s):
        mn, mx = s.min(), s.max()
        return pd.Series(50, index=s.index) if mx == mn else ((s - mn) / (mx - mn) * 100)

    med = df.groupby('code_commune')['prix_m2'].transform('median')
    s_prix = norm(((med - df['prix_m2']) / med.replace(0, np.nan)).clip(0).fillna(0))
    s_vol  = norm(df.groupby('code_commune')['prix_m2'].transform('count').astype(float))
    cutoff = pd.Timestamp.now() - pd.DateOffset(months=12)
    s_dyn  = norm((df['date_mutation'] >= cutoff).astype(float))

    return (s_prix * w_prix + s_vol * w_vol + s_dyn * w_dyn).round(0).clip(0, 100).astype(int)


# ─── CRM (session state) ──────────────────────────────────────────────────────

crm_db.init_crm()

STAGES = ["Détecté", "Contacté", "Qualifié", "Proposition", "Closing"]
STAGE_COLORS = {"Détecté":"#e5e5e5","Contacté":"#fff3cd","Qualifié":"#cfe2ff",
                "Proposition":"#d1ecf1","Closing":"#d4edda"}

# CRM functions — appels directs via crm_db
def add_contact(nom, email="", tel="", type_contact="Autre", notes=""):
    return crm_db.add_contact(nom, email, tel, type_contact, notes)

def add_opportunite(contact_id, titre, adresse, type_bien, surface, prix, prix_m2, score, source="DVF"):
    return crm_db.add_opportunite(contact_id, titre, adresse, type_bien, surface, prix, prix_m2, score, source)

def add_activite(opp_id, type_activite, notes="", date_str=None, statut="À faire"):
    return crm_db.add_activite(opp_id, type_activite, notes, date_str, statut)


# ─── SIDEBAR ──────────────────────────────────────────────────────────────────

DEPTS_COUVERTS = [
    "01", "02", "03", "04", "05", "06", "07", "08", "09", "10",
    "11", "12", "13", "14", "15", "16", "17", "18", "19",
    "21", "22", "23", "24", "25", "26", "27", "28", "29",
    "2A", "2B",
    "30", "31", "32", "33", "34", "35", "36", "37", "38", "39",
    "40", "41", "42", "43", "44", "45", "46", "47", "48", "49",
    "50", "51", "52", "53", "54", "55", "56", "57", "58", "59",
    "60", "61", "62", "63", "64", "65", "66", "67", "68", "69",
    "70", "71", "72", "73", "74", "75", "76", "77", "78", "79",
    "80", "81", "82", "83", "84", "85", "86", "87", "88", "89",
    "90", "91", "92", "93", "94", "95",
    "971", "972", "973", "974", "976",
]

DEPT_NOMS = {
    "01": "Ain", "02": "Aisne", "03": "Allier", "04": "Alpes-de-Haute-Provence",
    "05": "Hautes-Alpes", "06": "Alpes-Maritimes", "07": "Ardèche", "08": "Ardennes",
    "09": "Ariège", "10": "Aube", "11": "Aude", "12": "Aveyron",
    "13": "Bouches-du-Rhône", "14": "Calvados", "15": "Cantal", "16": "Charente",
    "17": "Charente-Maritime", "18": "Cher", "19": "Corrèze",
    "21": "Côte-d'Or", "22": "Côtes-d'Armor", "23": "Creuse",
    "24": "Dordogne", "25": "Doubs", "26": "Drôme", "27": "Eure",
    "28": "Eure-et-Loir", "29": "Finistère", "2A": "Corse-du-Sud", "2B": "Haute-Corse",
    "30": "Gard", "31": "Haute-Garonne", "32": "Gers", "33": "Gironde",
    "34": "Hérault", "35": "Ille-et-Vilaine", "36": "Indre", "37": "Indre-et-Loire",
    "38": "Isère", "39": "Jura", "40": "Landes", "41": "Loir-et-Cher",
    "42": "Loire", "43": "Haute-Loire", "44": "Loire-Atlantique", "45": "Loiret",
    "46": "Lot", "47": "Lot-et-Garonne", "48": "Lozère", "49": "Maine-et-Loire",
    "50": "Manche", "51": "Marne", "52": "Haute-Marne", "53": "Mayenne",
    "54": "Meurthe-et-Moselle", "55": "Meuse", "56": "Morbihan", "57": "Moselle",
    "58": "Nièvre", "59": "Nord", "60": "Oise", "61": "Orne",
    "62": "Pas-de-Calais", "63": "Puy-de-Dôme", "64": "Pyrénées-Atlantiques",
    "65": "Hautes-Pyrénées", "66": "Pyrénées-Orientales", "67": "Bas-Rhin",
    "68": "Haut-Rhin", "69": "Rhône", "70": "Haute-Saône", "71": "Saône-et-Loire",
    "72": "Sarthe", "73": "Savoie", "74": "Haute-Savoie", "75": "Paris",
    "76": "Seine-Maritime", "77": "Seine-et-Marne", "78": "Yvelines",
    "79": "Deux-Sèvres", "80": "Somme", "81": "Tarn", "82": "Tarn-et-Garonne",
    "83": "Var", "84": "Vaucluse", "85": "Vendée", "86": "Vienne",
    "87": "Haute-Vienne", "88": "Vosges", "89": "Yonne", "90": "Territoire de Belfort",
    "91": "Essonne", "92": "Hauts-de-Seine", "93": "Seine-Saint-Denis",
    "94": "Val-de-Marne", "95": "Val-d'Oise",
    "971": "Guadeloupe", "972": "Martinique", "973": "Guyane",
    "974": "La Réunion", "976": "Mayotte",
}

with st.sidebar:
    st.markdown(
        '<div style="text-align:center;padding:.5rem 0">'
        '<span style="font-size:1.4rem;font-weight:700">SAHAR</span>'
        '<span style="font-size:.8rem;color:#666;margin-left:.3rem">Conseil</span>'
        '</div>', unsafe_allow_html=True
    )
    st.markdown("---")

    # ── Localisation ──
    with st.expander("📍 Localisation", expanded=True):
        dept = st.selectbox(
            "Département", options=DEPTS_COUVERTS,
            index=DEPTS_COUVERTS.index("33"),
            format_func=lambda d: f"{d} — {DEPT_NOMS.get(d, d)}",
            help="101 départements couverts (DVF + DPE)",
        )
        # Filtre ville — rempli dynamiquement après chargement DVF
        ville_filter = st.text_input(
            "Filtrer par ville", "", key="ville_filter",
            help="Tapez le début du nom de ville pour filtrer (ex: Bordeaux)",
        )

    # ── Période ──
    with st.expander("📅 Période", expanded=True):
        periodes_presets = {
            "12 mois": 12, "24 mois": 24, "36 mois": 36,
            "5 ans": 60, "Personnalisé": 0
        }
        preset_periode = st.radio("Période d'analyse", list(periodes_presets.keys()),
                                   index=2, horizontal=True)
        if preset_periode == "Personnalisé":
            mois_periode = st.slider("Nombre de mois", 3, 84, 36, 3)
        else:
            mois_periode = periodes_presets[preset_periode]

        # Période de comparaison
        compare_label = st.radio("Comparer avec", ["N-1 (même durée)", "Période précédente", "Pas de comparaison"],
                                  index=0, horizontal=True)

    # ── Filtres biens ──
    with st.expander("🏠 Filtres biens", expanded=False):
        types_bien = st.multiselect("Type de bien",
                                     ["Appartement", "Maison"],
                                     default=["Appartement", "Maison"])

        col_s, col_e = st.columns(2)
        with col_s: surf_min = st.number_input("Surface min m²", 5, 500, 20, 5)
        with col_e: surf_max = st.number_input("Surface max m²", 20, 1000, 300, 10)

        col_p1, col_p2 = st.columns(2)
        with col_p1: prix_min = st.number_input("Prix min €/m²", 0, 25000, 500, 250)
        with col_p2: prix_max = st.number_input("Prix max €/m²", 500, 50000, 15000, 500)

        nb_pieces_min = st.number_input("Pièces minimum", 0, 10, 0, 1)
        score_min = st.slider("Score opportunité min", 0, 100, 0, 5)

    # ── Pondération scoring (mode avancé) ──
    with st.expander("⚙️ Scoring avancé", expanded=False):
        w_prix = st.slider("Sous-valorisation", 0.0, 1.0, 0.40, 0.05)
        w_vol  = st.slider("Volume marché",     0.0, 1.0, 0.30, 0.05)
        w_dyn  = st.slider("Dynamisme",         0.0, 1.0, 0.30, 0.05)
        total_w = w_prix + w_vol + w_dyn
        if abs(total_w - 1.0) > 0.01:
            st.warning(f"Somme poids = {total_w:.2f} ≠ 1")

    st.markdown("---")
    col_btn1, col_btn2 = st.columns(2)
    with col_btn1:
        if st.button("🔄 Cache", use_container_width=True):
            st.cache_data.clear()
            st.rerun()
    with col_btn2:
        if st.button("🔓 Déco", use_container_width=True):
            st.session_state.pop("auth_ok", None)
            st.rerun()


# ─── CHARGEMENT + FILTRAGE ───────────────────────────────────────────────────

with st.spinner("Chargement des données DVF..."):
    df_raw = load_dvf(dept)

now = pd.Timestamp.now()
cutoff = now - pd.DateOffset(months=mois_periode)

# Période courante
mask = (
    (df_raw['date_mutation'] >= cutoff) &
    (df_raw['type_local'].isin(types_bien)) &
    (df_raw['surface_utile'].between(surf_min, surf_max)) &
    (df_raw['prix_m2'].between(prix_min, prix_max))
)
if nb_pieces_min > 0 and 'nombre_pieces_principales' in df_raw.columns:
    mask = mask & (df_raw['nombre_pieces_principales'] >= nb_pieces_min)
# Filtre ville
if ville_filter:
    mask = mask & (df_raw['nom_commune'].str.contains(ville_filter, case=False, na=False))
df = df_raw[mask].copy()

# Selectbox commune (dynamique basée sur les données filtrées)
_communes_dispo = sorted(df_raw[df_raw['date_mutation'] >= cutoff]['nom_commune'].dropna().unique())
with st.sidebar:
    with st.expander("📍 Localisation", expanded=False):
        if _communes_dispo:
            _commune_choisie = st.selectbox(
                "Ou choisir une commune", ["Toutes"] + _communes_dispo,
                key="commune_select",
            )
            if _commune_choisie != "Toutes":
                df = df[df['nom_commune'] == _commune_choisie].copy()

# Période de comparaison (N-1)
if compare_label == "N-1 (même durée)":
    cutoff_comp_start = cutoff - pd.DateOffset(months=mois_periode)
    cutoff_comp_end = cutoff
elif compare_label == "Période précédente":
    cutoff_comp_start = cutoff - pd.DateOffset(months=mois_periode)
    cutoff_comp_end = cutoff
else:
    cutoff_comp_start = None
    cutoff_comp_end = None

if cutoff_comp_start is not None:
    mask_comp = (
        (df_raw['date_mutation'] >= cutoff_comp_start) &
        (df_raw['date_mutation'] < cutoff_comp_end) &
        (df_raw['type_local'].isin(types_bien)) &
        (df_raw['surface_utile'].between(surf_min, surf_max)) &
        (df_raw['prix_m2'].between(prix_min, prix_max))
    )
    if nb_pieces_min > 0 and 'nombre_pieces_principales' in df_raw.columns:
        mask_comp = mask_comp & (df_raw['nombre_pieces_principales'] >= nb_pieces_min)
    df_comp = df_raw[mask_comp].copy()
else:
    df_comp = pd.DataFrame()

if df.empty:
    st.warning("Aucun résultat avec ces filtres. Élargissez les critères.")
    st.stop()

# Scoring
tw = w_prix + w_vol + w_dyn
df['score'] = score_opportunite(df,
    w_prix/tw if tw > 0 else 0.4,
    w_vol/tw  if tw > 0 else 0.3,
    w_dyn/tw  if tw > 0 else 0.3)

df = df[df['score'] >= score_min]
if df.empty:
    st.warning("Aucun bien avec ce score minimum.")
    st.stop()


# ─── HEADER ───────────────────────────────────────────────────────────────────

# Helper deltas
def _delta(val_now, val_prev):
    """Calcule un delta en % entre deux valeurs."""
    if val_prev and val_prev > 0 and val_now is not None:
        return round((val_now - val_prev) / val_prev * 100, 1)
    return None

def _fmt_delta(d):
    return f"{d:+.1f}%" if d is not None else None

# Métriques période courante
_prix_med = df['prix_m2'].median()
_surface_med = df['surface_utile'].median()
_prix_total_med = df['valeur_fonciere'].median()
nb_opps = (df['score'] >= 70).sum()
_nb_communes = df['nom_commune'].nunique()
_nb_mois = max(df['date_mutation'].dt.to_period('M').nunique(), 1)
_vol_mensuel = round(len(df) / _nb_mois)
_taux_opps = round(nb_opps / len(df) * 100, 1) if len(df) > 0 else 0
_q1 = df['prix_m2'].quantile(0.25)
_q3 = df['prix_m2'].quantile(0.75)
_liquidite = round(len(df) / max(_nb_communes, 1), 1)

# Métriques période comparaison (deltas)
if not df_comp.empty:
    _d_tx = _delta(len(df), len(df_comp))
    _d_prix = _delta(_prix_med, df_comp['prix_m2'].median())
    _d_surface = _delta(_surface_med, df_comp['surface_utile'].median())
    _d_total = _delta(_prix_total_med, df_comp['valeur_fonciere'].median())
    _d_vol = _delta(_vol_mensuel, round(len(df_comp) / max(df_comp['date_mutation'].dt.to_period('M').nunique(), 1)))
    _comp_label = f"vs {'N-1' if compare_label.startswith('N-1') else 'période préc.'}"
else:
    _d_tx = _d_prix = _d_surface = _d_total = _d_vol = None
    _comp_label = ""

# ── Titre + période ──
_date_debut = df['date_mutation'].min().strftime("%d/%m/%Y") if len(df) > 0 else ""
_date_fin = df['date_mutation'].max().strftime("%d/%m/%Y") if len(df) > 0 else ""
st.markdown(
    f'<div style="display:flex;align-items:baseline;gap:1rem;margin-bottom:.5rem">'
    f'<span style="font-size:1.6rem;font-weight:700">Dept {dept}</span>'
    f'<span style="color:#888;font-size:.85rem">{_date_debut} → {_date_fin} — '
    f'{len(df):,} transactions {_comp_label}</span>'
    f'</div>', unsafe_allow_html=True
)

# ── KPIs avec deltas ──
c1, c2, c3, c4, c5, c6 = st.columns(6)
c1.metric("Transactions", f"{len(df):,}", _fmt_delta(_d_tx))
c2.metric("Prix médian €/m²", f"{_prix_med:,.0f}", _fmt_delta(_d_prix))
c3.metric("Surface médiane", f"{_surface_med:.0f} m²", _fmt_delta(_d_surface))
c4.metric("Prix médian bien", f"{_prix_total_med/1000:.0f} k€", _fmt_delta(_d_total))
c5.metric("Volume /mois", f"{_vol_mensuel}", _fmt_delta(_d_vol))
c6.metric("Opportunités ≥70", f"{nb_opps} ({_taux_opps}%)")

# ── Mini-barre contextuelle ──
_badges = [
    f"<b>{_nb_communes}</b> communes",
    f"<b>{_liquidite}</b> tx/commune",
    f"Q1–Q3 : <b>{_q1:,.0f}–{_q3:,.0f}</b> €/m²",
]
_crm_opps = len(st.session_state.crm_opportunites)
if _crm_opps > 0:
    _crm_closing = len([o for o in st.session_state.crm_opportunites if o.get("stage") == "Closing"])
    _crm_val = sum(o.get("prix", 0) for o in st.session_state.crm_opportunites)
    _badges.append(f"CRM: <b>{_crm_opps}</b> opps / <b>{_crm_closing}</b> closing / <b>{_crm_val/1000:.0f}k€</b>")

st.markdown(
    '<div style="display:flex;flex-wrap:wrap;gap:.8rem;padding:.4rem .8rem;'
    'background:#f8f9fa;border-radius:6px;font-size:.78rem;color:#555;margin:.3rem 0">'
    + ''.join(f'<span>{b}</span>' for b in _badges)
    + '</div>', unsafe_allow_html=True
)
st.markdown("")


# ─── ONGLETS PRINCIPAUX ───────────────────────────────────────────────────────

tab_data, tab_carte, tab_quartiers, tab_stats, tab_marche, tab_valo, tab_dpe, tab_crm, tab_pilotage, tab_export = st.tabs([
    "📋 Transactions",
    "🗺️ Carte",
    "🏘️ Quartiers",
    "📊 Statistiques",
    "🏙️ Marché par commune",
    "💰 Valorisation",
    "🔥 DPE × DVF",
    "💼 CRM Pipeline",
    "🎯 Pilotage",
    "📥 Export",
])


# ── TAB 1 : TRANSACTIONS ──────────────────────────────────────────────────────

with tab_data:
    col_left, col_right = st.columns([3, 1])

    with col_right:
        st.markdown("**Filtres rapides**")
        communes = ["Toutes"] + sorted(df['nom_commune'].dropna().unique().tolist())
        commune_sel = st.selectbox("Commune", communes)
        tri_col = st.selectbox("Trier par", ["score","prix_m2","valeur_fonciere","date_mutation","surface_utile"])
        tri_asc = st.toggle("Croissant", False)
        nb_lignes = st.slider("Nb lignes", 10, 500, 50, 10)

    with col_left:
        df_affich = df.copy()
        if commune_sel != "Toutes":
            df_affich = df_affich[df_affich['nom_commune'] == commune_sel]

        df_affich = df_affich.sort_values(tri_col, ascending=tri_asc)

        # Colonnes à afficher avec noms lisibles
        df_table = df_affich.head(nb_lignes)[[
            'score','date_mutation','nom_commune','adresse','code_postal',
            'type_local','surface_utile','nombre_pieces_principales',
            'surface_terrain','valeur_fonciere','prix_m2','id_parcelle'
        ]].copy()

        df_table['date_mutation'] = df_table['date_mutation'].dt.strftime('%d/%m/%Y')
        df_table['surface_terrain'] = df_table['surface_terrain'].fillna(0).astype(int)
        df_table['nombre_pieces_principales'] = df_table['nombre_pieces_principales'].fillna(0).astype(int)

        df_table = df_table.rename(columns={
            'score':'Score',
            'date_mutation':'Date vente',
            'nom_commune':'Commune',
            'adresse':'Adresse',
            'code_postal':'CP',
            'type_local':'Type',
            'surface_utile':'Surface m²',
            'nombre_pieces_principales':'Pièces',
            'surface_terrain':'Terrain m²',
            'valeur_fonciere':'Prix €',
            'prix_m2':'€/m²',
            'id_parcelle':'Parcelle',
        })

        st.dataframe(
            df_table,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Score": st.column_config.ProgressColumn("Score", min_value=0, max_value=100, format="%d"),
                "Prix €": st.column_config.NumberColumn("Prix €", format="%d €"),
                "€/m²": st.column_config.NumberColumn("€/m²", format="%d €"),
                "Surface m²": st.column_config.NumberColumn("Surface m²", format="%.0f"),
                "Terrain m²": st.column_config.NumberColumn("Terrain m²", format="%d"),
            }
        )
        st.caption(f"{len(df_affich):,} transactions • Affichage des {nb_lignes} premières")

        # ── CARD ANALYSE & RECOMMANDATION PAR BIEN ────────────────────
        st.markdown("---")
        st.markdown("**🔍 Analyse détaillée d'un bien**")
        if len(df_affich) > 0:
            bien_idx = st.selectbox(
                "Sélectionner un bien à analyser",
                range(min(nb_lignes, len(df_affich))),
                format_func=lambda i: (
                    f"{df_affich.iloc[i]['adresse']} — {df_affich.iloc[i]['nom_commune']} "
                    f"({df_affich.iloc[i]['prix_m2']:.0f} €/m²)"
                ) if i < len(df_affich) else "",
                key="analyse_bien_select"
            )
            bien = df_affich.iloc[bien_idx]

            # Calcul des indicateurs pour la recommandation
            _prix_commune_med = df[df['nom_commune'] == bien['nom_commune']]['prix_m2'].median()
            _ecart_prix = ((bien['prix_m2'] - _prix_commune_med) / _prix_commune_med * 100) if _prix_commune_med > 0 else 0
            _nb_tx_commune = len(df[df['nom_commune'] == bien['nom_commune']])
            _prix_dept_med = df['prix_m2'].median()
            _ecart_dept = ((bien['prix_m2'] - _prix_dept_med) / _prix_dept_med * 100) if _prix_dept_med > 0 else 0

            # Transactions récentes sur la même voie
            _voie_tx = df[df['adresse'].str.contains(
                bien.get('adresse_nom_voie', bien['adresse'].split()[-1]) if isinstance(bien.get('adresse_nom_voie', ''), str) else '',
                case=False, na=False
            )] if bien.get('adresse_nom_voie', '') else pd.DataFrame()
            _rotation_voie = len(_voie_tx)

            # Score et recommandation
            reco_points = []
            reco_score = bien['score']

            if _ecart_prix < -15:
                reco_points.append(("SOUS-ÉVALUÉ", f"{abs(_ecart_prix):.0f}% sous le prix médian commune", "#1D9E75"))
            elif _ecart_prix > 15:
                reco_points.append(("SURÉVALUÉ", f"{_ecart_prix:.0f}% au-dessus du prix médian commune", "#E24B4A"))
            else:
                reco_points.append(("PRIX MARCHÉ", f"Dans la fourchette commune (±{abs(_ecart_prix):.0f}%)", "#BA7517"))

            if _nb_tx_commune >= 20:
                reco_points.append(("MARCHÉ LIQUIDE", f"{_nb_tx_commune} transactions dans la commune", "#1D9E75"))
            elif _nb_tx_commune >= 5:
                reco_points.append(("MARCHÉ MODÉRÉ", f"{_nb_tx_commune} transactions", "#BA7517"))
            else:
                reco_points.append(("MARCHÉ PEU LIQUIDE", f"Seulement {_nb_tx_commune} transactions", "#E24B4A"))

            if bien['score'] >= 70:
                reco_points.append(("OPPORTUNITÉ FORTE", f"Score {bien['score']}/100", "#1D9E75"))
            elif bien['score'] >= 40:
                reco_points.append(("À SURVEILLER", f"Score {bien['score']}/100", "#BA7517"))

            # Générer la recommandation
            if bien['score'] >= 70 and _ecart_prix < -10:
                reco_text = "Bien sous-évalué dans un marché dynamique. Potentiel d'achat-revente ou de mise en location intéressant."
                reco_action = "ACHETER / INVESTIR"
                reco_color = "#1D9E75"
            elif bien['score'] >= 50 and _ecart_prix < 5:
                reco_text = "Bien à prix correct dans un marché actif. Bon potentiel si rénovation ou optimisation possible."
                reco_action = "ÉTUDIER"
                reco_color = "#BA7517"
            elif _ecart_prix > 20:
                reco_text = "Bien au-dessus du marché. Marge de négociation probable ou repositionnement nécessaire."
                reco_action = "NÉGOCIER"
                reco_color = "#E24B4A"
            else:
                reco_text = "Bien dans la moyenne du marché. Analyser le potentiel de valorisation spécifique."
                reco_action = "ANALYSER"
                reco_color = "#555"

            # Affichage Card
            badges_html = " ".join(
                f'<span style="background:{c};color:white;padding:2px 8px;border-radius:12px;font-size:.7rem;font-weight:600">{label}</span>'
                for label, detail, c in reco_points
            )
            details_html = "<br>".join(
                f'<span style="color:{c};font-weight:600">{label}</span> — {detail}'
                for label, detail, c in reco_points
            )

            _sv_url = ""
            if pd.notna(bien.get('latitude')) and pd.notna(bien.get('longitude')):
                _sv_url = f"https://www.google.com/maps/@?api=1&map_action=pano&viewpoint={bien['latitude']},{bien['longitude']}"
                _gm_url = f"https://www.google.com/maps/search/?api=1&query={bien['latitude']},{bien['longitude']}"

            st.markdown(f"""
            <div style="border:2px solid {reco_color};border-radius:12px;padding:16px 20px;background:#fafafa;margin:8px 0">
              <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px">
                <div>
                  <div style="font-size:1.1rem;font-weight:700">{bien['type_local']} — {bien['adresse']}</div>
                  <div style="font-size:.85rem;color:#666">{bien['nom_commune']} {bien['code_postal']}</div>
                </div>
                <div style="background:{reco_color};color:white;padding:6px 16px;border-radius:20px;font-weight:700;font-size:.85rem">
                  {reco_action}
                </div>
              </div>
              <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin:12px 0">
                <div style="text-align:center"><div style="font-size:1.3rem;font-weight:700">{bien['valeur_fonciere']:,.0f} €</div><div style="font-size:.7rem;color:#888">Prix vente</div></div>
                <div style="text-align:center"><div style="font-size:1.3rem;font-weight:700">{bien['prix_m2']:,.0f} €/m²</div><div style="font-size:.7rem;color:#888">Prix au m²</div></div>
                <div style="text-align:center"><div style="font-size:1.3rem;font-weight:700">{bien['surface_utile']:.0f} m²</div><div style="font-size:.7rem;color:#888">Surface</div></div>
                <div style="text-align:center"><div style="font-size:1.3rem;font-weight:700;color:{reco_color}">{bien['score']}/100</div><div style="font-size:.7rem;color:#888">Score</div></div>
              </div>
              <div style="margin:10px 0">{badges_html}</div>
              <div style="font-size:.82rem;margin:8px 0;line-height:1.6">{details_html}</div>
              <div style="background:{'#e8f5e9' if reco_color=='#1D9E75' else '#fff3e0' if reco_color=='#BA7517' else '#fce4ec'};padding:10px 14px;border-radius:8px;margin:8px 0">
                <div style="font-size:.85rem;font-weight:600;color:{reco_color}">💡 Recommandation</div>
                <div style="font-size:.82rem;color:#333">{reco_text}</div>
              </div>
              {'<div style="margin-top:8px;font-size:.8rem"><a href="' + _sv_url + '" target="_blank" style="color:#1a73e8;text-decoration:none">📍 Voir en Street View</a> &nbsp;·&nbsp; <a href="' + _gm_url + '" target="_blank" style="color:#1a73e8;text-decoration:none">🗺️ Google Maps</a></div>' if _sv_url else ''}
            </div>
            """, unsafe_allow_html=True)

        # ── AJOUT CRM : unitaire ou batch ─────────────────────────────
        st.markdown("---")
        st.markdown("**➕ Ajouter au CRM**")

        _crm_mode = st.radio("Mode", ["Ajout unitaire", "Ajout batch (top opportunités)"],
                              horizontal=True, label_visibility="collapsed")

        if _crm_mode == "Ajout unitaire":
            idx_sel = st.selectbox("Sélectionner une ligne",
                                   range(min(nb_lignes, len(df_affich))),
                                   format_func=lambda i: f"{df_affich.iloc[i]['adresse']} — {df_affich.iloc[i]['nom_commune']} ({df_affich.iloc[i]['prix_m2']:.0f} €/m²)" if i < len(df_affich) else "")
            if len(df_affich) > 0:
                row = df_affich.iloc[idx_sel]
                st.markdown(f"**{row['adresse']}, {row['nom_commune']} {row['code_postal']}**")
                st.markdown(f"{row['type_local']} — {row['surface_utile']:.0f} m² — {row['valeur_fonciere']:,.0f} € — {row['prix_m2']:.0f} €/m² — Score {row['score']}/100")

                contact_options = ["Créer nouveau contact"] + [f"{c['id']} — {c['nom']}" for c in st.session_state.crm_contacts]
                contact_choix = st.selectbox("Contact associé", contact_options)

                if contact_choix == "Créer nouveau contact":
                    nom_c = st.text_input("Nom")
                    email_c = st.text_input("Email")
                    tel_c = st.text_input("Téléphone")
                    type_c = st.selectbox("Type", ["Vendeur","Acheteur","Investisseur","Promoteur","Agent"])
                else:
                    contact_id = contact_choix.split(" — ")[0]

                if st.button("Ajouter au CRM", type="primary"):
                    if contact_choix == "Créer nouveau contact" and nom_c:
                        add_contact(nom_c, email_c, tel_c, type_c)
                        contact_id = st.session_state.crm_contacts[-1]['id']

                    if contact_choix != "Créer nouveau contact" or (contact_choix == "Créer nouveau contact" and nom_c):
                        add_opportunite(
                            contact_id,
                            titre=f"{row['type_local']} {row['adresse']}, {row['nom_commune']}",
                            adresse=f"{row['adresse']}, {row['nom_commune']} {row['code_postal']}",
                            type_bien=row['type_local'],
                            surface=float(row['surface_utile']),
                            prix=float(row['valeur_fonciere']),
                            prix_m2=float(row['prix_m2']),
                            score=int(row['score'])
                        )
                        st.success("Opportunité ajoutée au CRM !")
                        st.rerun()

        else:  # BATCH MODE
            _batch_score_min = st.slider("Score minimum pour batch", 40, 100, 70, 5, key="batch_score")
            _batch_max = st.slider("Nombre max à ajouter", 5, 50, 10, 5, key="batch_max")
            _batch_df = df_affich[df_affich['score'] >= _batch_score_min].nlargest(_batch_max, 'score')

            st.caption(f"{len(_batch_df)} opportunités score ≥ {_batch_score_min} prêtes à être ajoutées")

            if not _batch_df.empty:
                # Aperçu
                _batch_preview = _batch_df[['score','nom_commune','adresse','type_local','surface_utile','prix_m2','valeur_fonciere']].copy()
                _batch_preview = _batch_preview.rename(columns={
                    'score':'Score','nom_commune':'Commune','adresse':'Adresse',
                    'type_local':'Type','surface_utile':'m²','prix_m2':'€/m²','valeur_fonciere':'Prix'
                })
                st.dataframe(_batch_preview, use_container_width=True, hide_index=True,
                             column_config={
                                 "Score": st.column_config.ProgressColumn("Score",min_value=0,max_value=100,format="%d"),
                                 "Prix": st.column_config.NumberColumn(format="%d €"),
                                 "€/m²": st.column_config.NumberColumn(format="%d €"),
                             })

                # Contact batch
                _batch_contact = st.selectbox("Contact pour toutes les opportunités",
                    ["Auto (créer par commune)"] + [f"{c['id']} — {c['nom']}" for c in st.session_state.crm_contacts],
                    key="batch_contact")

                if st.button(f"🚀 Ajouter {len(_batch_df)} opportunités au CRM", type="primary", use_container_width=True):
                    _added = 0
                    for _, _row in _batch_df.iterrows():
                        # Vérifier doublon
                        _already = any(
                            o.get("adresse","").startswith(str(_row['adresse'])[:20])
                            for o in st.session_state.crm_opportunites
                        )
                        if _already:
                            continue

                        if _batch_contact == "Auto (créer par commune)":
                            _commune_name = str(_row['nom_commune'])
                            _existing = next((c for c in st.session_state.crm_contacts
                                              if _commune_name.lower() in c['nom'].lower()), None)
                            if _existing:
                                _cid = _existing['id']
                            else:
                                add_contact(f"Prospect {_commune_name}", type_contact="Vendeur")
                                _cid = st.session_state.crm_contacts[-1]['id']
                        else:
                            _cid = _batch_contact.split(" — ")[0]

                        add_opportunite(
                            _cid,
                            titre=f"{_row['type_local']} {_row['adresse']}, {_row['nom_commune']}",
                            adresse=f"{_row['adresse']}, {_row['nom_commune']} {_row.get('code_postal','')}",
                            type_bien=str(_row['type_local']),
                            surface=float(_row['surface_utile']),
                            prix=float(_row['valeur_fonciere']),
                            prix_m2=float(_row['prix_m2']),
                            score=int(_row['score'])
                        )
                        _added += 1

                    st.success(f"✅ {_added} opportunités ajoutées au CRM !")
                    if _added > 0:
                        st.rerun()
            else:
                st.info("Aucune opportunité avec ce score minimum.")


# ── TAB 2 : CARTE (améliorée — heatmap, couches, Street View) ────────────────

with tab_carte:
    df_carte = df.dropna(subset=['latitude','longitude']).copy()

    if df_carte.empty:
        st.info("Coordonnées GPS non disponibles. Utilisez les fichiers geo-dvf complets.")
    else:
        try:
            import folium
            from folium.plugins import HeatMap, MarkerCluster
            from streamlit_folium import st_folium

            # ── Contrôles carte ──
            col_c1, col_c2, col_c3 = st.columns(3)
            with col_c1:
                mode_carte = st.radio("Mode", ["Points", "Heatmap prix/m²", "Heatmap volume", "Clusters"], horizontal=True)
            with col_c2:
                max_points = st.slider("Nb points", 200, 5000, 1500, 100, key="map_pts")
            with col_c3:
                fond_carte = st.selectbox("Fond de carte", ["CartoDB positron", "OpenStreetMap", "CartoDB dark_matter"])

            df_map = df_carte.nlargest(max_points, 'score')
            lat_c, lon_c = df_map['latitude'].median(), df_map['longitude'].median()
            m = folium.Map(location=[lat_c, lon_c], zoom_start=12, tiles=fond_carte)

            def couleur_prix(prix):
                if prix >= 5000: return "#8B0000"
                if prix >= 3500: return "#E24B4A"
                if prix >= 2500: return "#BA7517"
                if prix >= 1500: return "#1D9E75"
                return "#0066CC"

            def couleur_score(score):
                if score >= 70: return "#1D9E75"
                if score >= 40: return "#BA7517"
                return "#E24B4A"

            if mode_carte == "Heatmap prix/m²":
                heat_data = [[r['latitude'], r['longitude'], r['prix_m2']] for _, r in df_map.iterrows()]
                HeatMap(heat_data, radius=18, blur=15, gradient={0.2:'blue',0.4:'lime',0.6:'yellow',0.8:'orange',1:'red'}).add_to(m)
                st.caption("Heatmap basée sur le prix au m² — rouge = plus cher, bleu = moins cher")

            elif mode_carte == "Heatmap volume":
                heat_data = [[r['latitude'], r['longitude']] for _, r in df_map.iterrows()]
                HeatMap(heat_data, radius=15, blur=12).add_to(m)
                st.caption("Heatmap basée sur la densité de transactions")

            elif mode_carte == "Clusters":
                mc = MarkerCluster()
                for _, row in df_map.iterrows():
                    sv_url = f"https://www.google.com/maps/@?api=1&map_action=pano&viewpoint={row['latitude']},{row['longitude']}"
                    popup = f"""
                    <div style='font-family:sans-serif;font-size:11px;min-width:220px'>
                      <b>{row['type_local']} — {row['adresse']}</b><br>
                      {row['nom_commune']} {row['code_postal']}<br>
                      <b>Prix :</b> {row['valeur_fonciere']:,.0f} € &nbsp; <b>Surface :</b> {row['surface_utile']:.0f} m²<br>
                      <b>Prix/m² :</b> {row['prix_m2']:.0f} € &nbsp;
                      <b style='color:{couleur_score(row["score"])}'>Score {row["score"]}</b><br>
                      <b>Date :</b> {row['date_mutation'].strftime('%d/%m/%Y') if pd.notna(row['date_mutation']) else '—'}<br>
                      <a href="{sv_url}" target="_blank" style="color:#1a73e8">📍 Street View</a>
                      &nbsp;·&nbsp;
                      <a href="https://www.google.com/maps/search/?api=1&query={row['latitude']},{row['longitude']}" target="_blank" style="color:#1a73e8">🗺️ Google Maps</a>
                    </div>"""
                    folium.Marker(
                        [row['latitude'], row['longitude']],
                        popup=folium.Popup(popup, max_width=280),
                        tooltip=f"{row['prix_m2']:.0f} €/m²",
                        icon=folium.Icon(color='green' if row['score']>=70 else 'orange' if row['score']>=40 else 'red', icon='home', prefix='fa')
                    ).add_to(mc)
                mc.add_to(m)

            else:  # Points (par défaut)
                for _, row in df_map.iterrows():
                    sv_url = f"https://www.google.com/maps/@?api=1&map_action=pano&viewpoint={row['latitude']},{row['longitude']}"
                    terrain = f" | Terrain {row['surface_terrain']:.0f}m²" if row.get('surface_terrain', 0) > 0 else ""
                    pieces = f" | {int(row['nombre_pieces_principales'])}p" if row.get('nombre_pieces_principales', 0) > 0 else ""
                    popup = f"""
                    <div style='font-family:sans-serif;font-size:11px;min-width:220px'>
                      <b>{row['type_local']} — {row['adresse']}</b><br>
                      {row['nom_commune']} {row['code_postal']}<br>
                      <b>Prix :</b> {row['valeur_fonciere']:,.0f} € &nbsp; <b>Surface :</b> {row['surface_utile']:.0f}m²{terrain}{pieces}<br>
                      <b>Prix/m² :</b> {row['prix_m2']:.0f} € &nbsp;
                      <b style='color:{couleur_score(row["score"])}'>Score {row["score"]}</b><br>
                      <b>Date :</b> {row['date_mutation'].strftime('%d/%m/%Y') if pd.notna(row['date_mutation']) else '—'}<br>
                      <a href="{sv_url}" target="_blank" style="color:#1a73e8">📍 Street View</a>
                      &nbsp;·&nbsp;
                      <a href="https://www.google.com/maps/search/?api=1&query={row['latitude']},{row['longitude']}" target="_blank" style="color:#1a73e8">🗺️ Google Maps</a>
                    </div>"""
                    folium.CircleMarker(
                        location=[row['latitude'], row['longitude']],
                        radius=6,
                        color=couleur_prix(row['prix_m2']),
                        fill=True, fill_color=couleur_prix(row['prix_m2']),
                        fill_opacity=0.8,
                        popup=folium.Popup(popup, max_width=280),
                        tooltip=f"{row['type_local']} — {row['prix_m2']:.0f} €/m² — Score {row['score']}"
                    ).add_to(m)

                # Légende prix
                col_leg = st.columns(5)
                with col_leg[0]: st.markdown('<span style="color:#0066CC">● < 1 500</span>', unsafe_allow_html=True)
                with col_leg[1]: st.markdown('<span style="color:#1D9E75">● 1 500–2 500</span>', unsafe_allow_html=True)
                with col_leg[2]: st.markdown('<span style="color:#BA7517">● 2 500–3 500</span>', unsafe_allow_html=True)
                with col_leg[3]: st.markdown('<span style="color:#E24B4A">● 3 500–5 000</span>', unsafe_allow_html=True)
                with col_leg[4]: st.markdown('<span style="color:#8B0000">● > 5 000 €/m²</span>', unsafe_allow_html=True)

            st_folium(m, height=560, use_container_width=True)
            st.caption(f"{len(df_map):,} points affichés — cliquez sur un point pour détails + Street View")

        except ImportError:
            st.warning("Installer folium et streamlit-folium pour activer la carte.")
            st.code("pip install folium streamlit-folium")


# ── TAB 3 : QUARTIERS (rotation IRIS, prix/m² par zone) ─────────────────────

with tab_quartiers:
    st.subheader("🏘️ Analyse par quartier — Rotation et prix")
    st.caption("Segmentation par micro-zone (voie/quartier), rotation immobilière et valorisation par îlot")

    df_q = df.dropna(subset=['latitude', 'longitude']).copy()

    if df_q.empty:
        st.info("Pas de données géolocalisées pour l'analyse quartier.")
    else:
        # Créer des micro-zones en arrondissant les coordonnées (≈ 100m)
        precision = 3  # 3 décimales ≈ 100m de résolution
        df_q['zone_lat'] = df_q['latitude'].round(precision)
        df_q['zone_lon'] = df_q['longitude'].round(precision)
        df_q['zone_id'] = df_q['zone_lat'].astype(str) + "," + df_q['zone_lon'].astype(str)

        # Agréger par zone
        zone_stats = df_q.groupby('zone_id').agg(
            nb_transactions=('valeur_fonciere', 'count'),
            prix_m2_median=('prix_m2', 'median'),
            prix_m2_moy=('prix_m2', 'mean'),
            prix_m2_min=('prix_m2', 'min'),
            prix_m2_max=('prix_m2', 'max'),
            volume_total=('valeur_fonciere', 'sum'),
            surface_moy=('surface_utile', 'mean'),
            score_moy=('score', 'mean'),
            lat=('latitude', 'median'),
            lon=('longitude', 'median'),
            commune=('nom_commune', 'first'),
        ).reset_index()

        # Calculer la rotation (nb transactions / an normalisé)
        nb_mois_periode = max(mois_periode, 1)
        zone_stats['rotation_an'] = (zone_stats['nb_transactions'] / nb_mois_periode * 12).round(1)
        zone_stats['ecart_prix'] = ((zone_stats['prix_m2_max'] - zone_stats['prix_m2_min']) / zone_stats['prix_m2_median'].replace(0, np.nan) * 100).round(1).fillna(0)

        # Filtrer zones avec assez de données
        min_tx = st.slider("Minimum transactions par zone", 2, 20, 3, 1, key="zone_min_tx")
        zone_filtre = zone_stats[zone_stats['nb_transactions'] >= min_tx].sort_values('rotation_an', ascending=False).copy()

        if zone_filtre.empty:
            st.info("Pas assez de données pour cette granularité. Baissez le minimum de transactions.")
        else:
            # KPIs
            zk1, zk2, zk3, zk4 = st.columns(4)
            zk1.metric("Zones actives", len(zone_filtre))
            zk2.metric("Rotation moy./an", f"{zone_filtre['rotation_an'].mean():.1f}")
            zk3.metric("Prix médian zone", f"{zone_filtre['prix_m2_median'].median():,.0f} €/m²")
            zk4.metric("Écart prix moyen", f"{zone_filtre['ecart_prix'].mean():.0f}%")

            st.markdown("---")

            col_map_q, col_table_q = st.columns([1, 1])

            with col_map_q:
                st.markdown("##### Carte des zones — couleur = prix/m²")
                try:
                    import folium
                    from streamlit_folium import st_folium

                    m_q = folium.Map(location=[zone_filtre['lat'].median(), zone_filtre['lon'].median()],
                                     zoom_start=12, tiles="CartoDB positron")
                    for _, z in zone_filtre.iterrows():
                        radius = max(4, min(15, z['nb_transactions'] * 1.5))
                        folium.CircleMarker(
                            [z['lat'], z['lon']], radius=radius,
                            color=couleur_prix(z['prix_m2_median']),
                            fill=True, fill_color=couleur_prix(z['prix_m2_median']),
                            fill_opacity=0.7,
                            tooltip=f"{z['commune']} — {z['prix_m2_median']:,.0f} €/m² — {z['nb_transactions']} tx — Rotation {z['rotation_an']}/an",
                            popup=folium.Popup(
                                f"<div style='font-size:11px'>"
                                f"<b>{z['commune']}</b><br>"
                                f"Prix médian: {z['prix_m2_median']:,.0f} €/m²<br>"
                                f"Transactions: {z['nb_transactions']}<br>"
                                f"Rotation: {z['rotation_an']}/an<br>"
                                f"Volume: {z['volume_total']:,.0f} €<br>"
                                f"<a href='https://www.google.com/maps/@?api=1&map_action=pano&viewpoint={z['lat']},{z['lon']}' target='_blank'>📍 Street View</a>"
                                f"</div>", max_width=250
                            ),
                        ).add_to(m_q)
                    st_folium(m_q, height=450, use_container_width=True)
                except ImportError:
                    st.warning("folium requis pour la carte quartiers")

            with col_table_q:
                st.markdown("##### Top zones par rotation")
                display_z = zone_filtre.head(30)[['commune', 'nb_transactions', 'rotation_an',
                                                    'prix_m2_median', 'prix_m2_min', 'prix_m2_max',
                                                    'ecart_prix', 'volume_total', 'score_moy']].copy()
                display_z.columns = ['Commune', 'Transactions', 'Rotation/an', 'Prix médian €/m²',
                                     'Prix min', 'Prix max', 'Écart %', 'Volume €', 'Score moy']
                st.dataframe(display_z, use_container_width=True, height=450,
                             column_config={
                                 "Score moy": st.column_config.ProgressColumn(min_value=0, max_value=100, format="%.0f"),
                             })

            # ── Graphique rotation vs prix ──
            st.markdown("---")
            st.markdown("##### Rotation vs Prix par zone")
            try:
                import plotly.express as px
                fig_rp = px.scatter(
                    zone_filtre, x='prix_m2_median', y='rotation_an',
                    size='nb_transactions', color='score_moy',
                    color_continuous_scale='RdYlGn', hover_name='commune',
                    hover_data={'volume_total': ':,.0f', 'ecart_prix': ':.1f'},
                    labels={'prix_m2_median': 'Prix médian €/m²', 'rotation_an': 'Rotation annuelle',
                            'nb_transactions': 'Volume', 'score_moy': 'Score'},
                    title="Zones à forte rotation + prix attractif = opportunités"
                )
                fig_rp.update_layout(height=400)
                st.plotly_chart(fig_rp, use_container_width=True)
            except ImportError:
                pass


# ── TAB 4 : STATISTIQUES ──────────────────────────────────────────────────────

with tab_stats:
    try:
        import plotly.express as px

        col_g1, col_g2 = st.columns(2)

        with col_g1:
            df_mois = df.groupby('mois')['prix_m2'].median().reset_index()
            fig1 = px.line(df_mois, x='mois', y='prix_m2',
                           title='Évolution prix médian €/m²',
                           labels={'mois':'Mois','prix_m2':'€/m²'},
                           color_discrete_sequence=['#185FA5'])
            fig1.update_layout(plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
            st.plotly_chart(fig1, use_container_width=True)

        with col_g2:
            fig2 = px.histogram(df, x='score', nbins=20,
                                title='Distribution des scores',
                                color_discrete_sequence=['#1D9E75'])
            fig2.add_vline(x=70, line_dash='dash', line_color='#1D9E75')
            fig2.add_vline(x=40, line_dash='dash', line_color='#BA7517')
            fig2.update_layout(plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
            st.plotly_chart(fig2, use_container_width=True)

        col_g3, col_g4 = st.columns(2)

        with col_g3:
            fig3 = px.box(df, x='type_local', y='prix_m2',
                          title='Prix €/m² par type',
                          color='type_local',
                          color_discrete_map={'Appartement':'#185FA5','Maison':'#1D9E75'})
            fig3.update_layout(plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', showlegend=False)
            st.plotly_chart(fig3, use_container_width=True)

        with col_g4:
            top_communes = df.groupby('nom_commune').agg(
                nb=('prix_m2','count'), mediane=('prix_m2','median')
            ).nlargest(15,'nb').reset_index()
            fig4 = px.bar(top_communes, x='nb', y='nom_commune', orientation='h',
                          title='Top 15 communes par volume',
                          color='mediane', color_continuous_scale='Blues',
                          labels={'nb':'Transactions','nom_commune':'','mediane':'Médiane €/m²'})
            fig4.update_layout(plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
            st.plotly_chart(fig4, use_container_width=True)

    except ImportError:
        st.info("Installer plotly pour les graphiques : pip install plotly")


# ── TAB 4 : MARCHÉ PAR COMMUNE ───────────────────────────────────────────────

with tab_marche:
    st.markdown("### 🏙️ Scoring de tension marché par commune")
    st.caption(
        "Chaque commune est scorée sur 5 indicateurs : volume de transactions, "
        "évolution des prix sur 12 mois, ratio de tension offre/demande. "
        "Score 0–100 — plus c'est haut, plus le marché est dynamique."
    )

    with st.spinner("Calcul du scoring communes..."):
        df_comm = compute_score_commune(df)

    if df_comm.empty:
        st.warning("Pas assez de données pour scorer les communes (minimum 5 transactions/commune requises).")
    else:
        # ── Filtres rapides ───────────────────────────────────────────
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            signal_filter = st.multiselect(
                "Signal marché",
                options=df_comm["Signal marché"].unique().tolist(),
                default=df_comm["Signal marché"].unique().tolist()
            )
        with col_f2:
            score_min_c = st.slider("Score minimum", 0, 100, 0, 5, key="score_comm_min")
        with col_f3:
            nb_communes = st.slider("Nb communes affichées", 5, 100, 30, 5)

        df_comm_filtre = df_comm[
            (df_comm["Signal marché"].isin(signal_filter)) &
            (df_comm["Score marché"] >= score_min_c)
        ].head(nb_communes)

        # ── Métriques résumé ─────────────────────────────────────────
        m1, m2, m3, m4 = st.columns(4)
        with m1:
            st.metric("Communes analysées", len(df_comm))
        with m2:
            opps_fortes = (df_comm["Score marché"] >= 70).sum()
            st.metric("Marchés score ≥ 70", opps_fortes)
        with m3:
            vendeur = df_comm["Signal marché"].str.contains("vendeur", case=False).sum()
            st.metric("Marchés vendeurs", vendeur)
        with m4:
            acheteur = df_comm["Signal marché"].str.contains("Opportunité", case=False).sum()
            st.metric("Opportunités acheteur", acheteur)

        st.markdown("---")

        # ── Tableau scoring ───────────────────────────────────────────
        cols_display = [
            "Commune", "Score marché", "Signal marché",
            "Prix médian €/m²", "Évolution 12m (%)",
            "Transactions 12m", "Ratio tension",
        ]
        df_display = df_comm_filtre[[c for c in cols_display if c in df_comm_filtre.columns]].copy()

        st.dataframe(
            df_display,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Score marché": st.column_config.ProgressColumn(
                    "Score marché", min_value=0, max_value=100, format="%d"
                ),
                "Prix médian €/m²": st.column_config.NumberColumn(
                    "Prix médian €/m²", format="%d €"
                ),
                "Évolution 12m (%)": st.column_config.NumberColumn(
                    "Évolution 12m (%)", format="%.1f%%"
                ),
            }
        )

        st.markdown("---")

        # ── Graphique top communes ────────────────────────────────────
        try:
            import plotly.express as px
            top_graph = df_comm.head(20).copy()
            fig_comm_score = px.bar(
                top_graph,
                x="Score marché",
                y="Commune",
                orientation="h",
                color="Score marché",
                color_continuous_scale=["#E24B4A", "#BA7517", "#1D9E75"],
                title="Top 20 communes — Score de tension marché",
                labels={"Score marché": "Score", "Commune": ""},
                text="Score marché"
            )
            fig_comm_score.update_layout(
                height=520,
                plot_bgcolor="rgba(0,0,0,0)",
                paper_bgcolor="rgba(0,0,0,0)",
                coloraxis_showscale=False,
                yaxis={"categoryorder": "total ascending"},
                margin=dict(l=0, r=0, t=35, b=0)
            )
            fig_comm_score.update_traces(textposition="outside")
            st.plotly_chart(fig_comm_score, use_container_width=True)
        except ImportError:
            pass

        st.markdown("---")

        # ── Génération rapport PDF par commune ────────────────────────
        st.markdown("### 📄 Rapport PDF par commune")
        st.caption("Exportez un rapport A4 professionnel pour n'importe quelle commune — à envoyer à un client.")

        communes_dispo = df_comm["Commune"].tolist()
        commune_pdf = st.selectbox(
            "Choisir une commune",
            communes_dispo,
            key="commune_pdf_select"
        )

        if commune_pdf:
            row_pdf = df_comm[df_comm["Commune"] == commune_pdf].iloc[0].to_dict()
            df_commune_trans = df[df["nom_commune"] == commune_pdf].copy()
            nb_dept = len(df[df["date_mutation"] >= (pd.Timestamp.now() - pd.DateOffset(months=12))])

            col_prev, col_dl = st.columns([2, 1])

            with col_prev:
                score_v = row_pdf.get("Score marché", 0)
                evol_v  = row_pdf.get("Évolution 12m (%)", None)
                signal_v = row_pdf.get("Signal marché", "—")
                prix_v  = row_pdf.get("Prix médian €/m²", 0)
                vol_v   = row_pdf.get("Transactions 12m", 0)
                evol_str = f"{evol_v:+.1f}%" if evol_v is not None and str(evol_v) != "nan" else "N/D"

                score_color = "#1D9E75" if score_v >= 70 else "#BA7517" if score_v >= 40 else "#E24B4A"

                st.markdown(
                    f"""
                    <div style="border:1px solid #e5e5e5;border-radius:8px;padding:16px 20px;background:#fafafa">
                      <div style="font-size:1.1rem;font-weight:700;color:#185FA5;margin-bottom:10px">
                        Aperçu — {commune_pdf}
                      </div>
                      <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:10px">
                        <div style="text-align:center">
                          <div style="font-size:1.6rem;font-weight:700;color:{score_color}">{score_v}/100</div>
                          <div style="font-size:.75rem;color:#888">Score marché</div>
                        </div>
                        <div style="text-align:center">
                          <div style="font-size:1.6rem;font-weight:700;color:#185FA5">{prix_v:,.0f} €</div>
                          <div style="font-size:.75rem;color:#888">Prix médian €/m²</div>
                        </div>
                        <div style="text-align:center">
                          <div style="font-size:1.6rem;font-weight:700;color:#185FA5">{evol_str}</div>
                          <div style="font-size:.75rem;color:#888">Évolution 12m</div>
                        </div>
                      </div>
                      <div style="font-size:.85rem;color:#444">
                        {signal_v} &nbsp;·&nbsp; {int(vol_v)} transactions sur 12 mois
                      </div>
                    </div>
                    """,
                    unsafe_allow_html=True
                )

            with col_dl:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("🖨️ Générer le rapport PDF", type="primary", use_container_width=True):
                    with st.spinner("Génération du rapport..."):
                        try:
                            pdf_bytes = generer_rapport_commune(
                                commune=commune_pdf,
                                dept=dept,
                                row_commune=row_pdf,
                                df_transactions=df_commune_trans,
                                nb_total_dept=nb_dept
                            )
                            st.download_button(
                                label="📥 Télécharger le PDF",
                                data=pdf_bytes,
                                file_name=f"sahar_analyse_{commune_pdf.replace(' ', '_')}_{pd.Timestamp.now().strftime('%Y%m%d')}.pdf",
                                mime="application/pdf",
                                use_container_width=True,
                                type="primary"
                            )
                            st.success("✅ Rapport généré !")
                        except Exception as e:
                            st.error(f"Erreur génération PDF : {e}")


# ── TAB 6 : VALORISATION — Estimation prix par adresse (comparables DVF) ─────

with tab_valo:
    st.subheader("💰 Valorisation marché par adresse")
    st.caption("Estimation fiable basée sur les transactions DVF comparables — méthode des comparables ajustée")

    df_valo = df.dropna(subset=['latitude', 'longitude']).copy()

    if df_valo.empty:
        st.info("Pas de données géolocalisées pour la valorisation.")
    else:
        col_v1, col_v2 = st.columns([1, 1])

        with col_v1:
            st.markdown("##### Rechercher une adresse")
            adresse_recherche = st.text_input("Adresse ou nom de voie", "", key="valo_addr",
                                               help="Ex: Avenue des Champs Elysées, Rue de Rivoli...")
            type_recherche = st.radio("Type de bien", ["Appartement", "Maison", "Tous"], horizontal=True, key="valo_type")
            rayon_m = st.slider("Rayon de recherche (m)", 100, 2000, 500, 50, key="valo_rayon")

        with col_v2:
            st.markdown("##### Paramètres")
            surface_cible = st.number_input("Surface estimée (m²)", 10, 500, 70, 5, key="valo_surface")
            nb_comparables = st.slider("Nb comparables min", 3, 30, 5, 1, key="valo_nb_comp")
            ponderer_date = st.checkbox("Pondérer par ancienneté", value=True, key="valo_pond_date")

        if adresse_recherche:
            # Rechercher les transactions correspondantes
            mask_addr = df_valo['adresse'].str.contains(adresse_recherche, case=False, na=False)
            if type_recherche != "Tous":
                mask_addr = mask_addr & (df_valo['type_local'] == type_recherche)
            df_match = df_valo[mask_addr].copy()

            if df_match.empty:
                st.warning(f"Aucune transaction trouvée pour '{adresse_recherche}'")

                # Tenter une recherche par proximité GPS si on a une adresse avec des résultats proches
                df_voie = df_valo[df_valo['adresse'].str.contains(adresse_recherche.split()[0], case=False, na=False)]
                if not df_voie.empty:
                    st.info(f"Transactions proches trouvées : {len(df_voie)} sur la même voie/zone")
                    df_match = df_voie.head(50)
            if not df_match.empty:
                # Point de référence = centroïde des transactions trouvées
                ref_lat = df_match['latitude'].median()
                ref_lon = df_match['longitude'].median()

                # Trouver les comparables dans le rayon
                from math import radians, cos, sin, asin, sqrt
                def haversine(lat1, lon1, lat2, lon2):
                    lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
                    dlat, dlon = lat2-lat1, lon2-lon1
                    a = sin(dlat/2)**2 + cos(lat1)*cos(lat2)*sin(dlon/2)**2
                    return 2 * asin(sqrt(a)) * 6371000

                df_valo['distance_m'] = df_valo.apply(
                    lambda r: haversine(ref_lat, ref_lon, r['latitude'], r['longitude']), axis=1
                )
                df_comp_valo = df_valo[df_valo['distance_m'] <= rayon_m].copy()
                if type_recherche != "Tous":
                    df_comp_valo = df_comp_valo[df_comp_valo['type_local'] == type_recherche]

                df_comp_valo = df_comp_valo.sort_values('distance_m')

                if len(df_comp_valo) < nb_comparables:
                    st.warning(f"Seulement {len(df_comp_valo)} comparables dans {rayon_m}m (min: {nb_comparables})")

                if not df_comp_valo.empty:
                    # Calcul pondéré
                    if ponderer_date:
                        jours = (pd.Timestamp.now() - df_comp_valo['date_mutation']).dt.days.clip(1)
                        df_comp_valo['poids'] = (1 / jours) * (1 / df_comp_valo['distance_m'].clip(10))
                    else:
                        df_comp_valo['poids'] = 1 / df_comp_valo['distance_m'].clip(10)

                    poids_total = df_comp_valo['poids'].sum()
                    prix_estime_m2 = (df_comp_valo['prix_m2'] * df_comp_valo['poids']).sum() / poids_total if poids_total > 0 else df_comp_valo['prix_m2'].median()
                    prix_estime_total = prix_estime_m2 * surface_cible

                    # Intervalles de confiance
                    q10 = df_comp_valo['prix_m2'].quantile(0.1)
                    q90 = df_comp_valo['prix_m2'].quantile(0.9)

                    st.markdown("---")
                    st.markdown("##### Estimation de valeur")

                    vc1, vc2, vc3, vc4 = st.columns(4)
                    vc1.metric("Prix estimé/m²", f"{prix_estime_m2:,.0f} €")
                    vc2.metric("Valeur totale", f"{prix_estime_total:,.0f} €")
                    vc3.metric("Fourchette basse", f"{q10 * surface_cible:,.0f} €")
                    vc4.metric("Fourchette haute", f"{q90 * surface_cible:,.0f} €")

                    vc5, vc6, vc7, vc8 = st.columns(4)
                    vc5.metric("Comparables", len(df_comp_valo))
                    vc6.metric("Rayon effectif", f"{df_comp_valo['distance_m'].max():.0f} m")
                    vc7.metric("Prix min/m²", f"{df_comp_valo['prix_m2'].min():,.0f} €")
                    vc8.metric("Prix max/m²", f"{df_comp_valo['prix_m2'].max():,.0f} €")

                    # Carte des comparables
                    st.markdown("##### Carte des comparables")
                    try:
                        import folium
                        from streamlit_folium import st_folium

                        m_v = folium.Map(location=[ref_lat, ref_lon], zoom_start=15, tiles="CartoDB positron")

                        # Point de référence
                        folium.Marker([ref_lat, ref_lon],
                                      icon=folium.Icon(color='red', icon='star', prefix='fa'),
                                      tooltip="Adresse recherchée").add_to(m_v)

                        # Cercle rayon
                        folium.Circle([ref_lat, ref_lon], radius=rayon_m,
                                      color='blue', fill=False, opacity=0.3).add_to(m_v)

                        for _, r in df_comp_valo.head(50).iterrows():
                            sv_url = f"https://www.google.com/maps/@?api=1&map_action=pano&viewpoint={r['latitude']},{r['longitude']}"
                            folium.CircleMarker(
                                [r['latitude'], r['longitude']], radius=5,
                                color=couleur_prix(r['prix_m2']),
                                fill=True, fill_color=couleur_prix(r['prix_m2']), fill_opacity=0.8,
                                tooltip=f"{r['prix_m2']:,.0f} €/m² — {r['distance_m']:.0f}m",
                                popup=folium.Popup(
                                    f"<div style='font-size:11px'>"
                                    f"<b>{r['adresse']}</b><br>"
                                    f"Prix: {r['valeur_fonciere']:,.0f} € — {r['surface_utile']:.0f}m²<br>"
                                    f"Prix/m²: {r['prix_m2']:,.0f} €<br>"
                                    f"Date: {r['date_mutation'].strftime('%d/%m/%Y') if pd.notna(r['date_mutation']) else '—'}<br>"
                                    f"Distance: {r['distance_m']:.0f}m<br>"
                                    f"<a href='{sv_url}' target='_blank'>📍 Street View</a>"
                                    f"</div>", max_width=250
                                ),
                            ).add_to(m_v)
                        st_folium(m_v, height=400, use_container_width=True)
                    except ImportError:
                        pass

                    # Tableau comparables
                    st.markdown("##### Détail des comparables")
                    show_comp = df_comp_valo.head(20)[['adresse', 'nom_commune', 'type_local',
                                                        'valeur_fonciere', 'surface_utile', 'prix_m2',
                                                        'date_mutation', 'distance_m']].copy()
                    show_comp.columns = ['Adresse', 'Commune', 'Type', 'Prix €', 'Surface m²',
                                         'Prix/m²', 'Date', 'Distance m']
                    st.dataframe(show_comp, use_container_width=True, height=350)


# ── TAB 7 : DPE × DVF — Croisement passoires thermiques / transactions ───────

with tab_dpe:
    st.subheader("🔥 Croisement DPE × DVF — Opportunités rénovation")
    st.caption("Identifie les communes avec forte activité immobilière ET forte concentration de passoires thermiques")

    # Charger les données DPE
    try:
        from shared.supabase_dpe import get_dpe_communes, get_dpe_logements, get_dpe_stats
        dpe_communes = get_dpe_communes(departement=dept)
        has_dpe = not dpe_communes.empty
    except Exception as e:
        has_dpe = False
        dpe_communes = pd.DataFrame()

    if not has_dpe:
        st.warning(f"Pas de données DPE disponibles pour le département {dept}. "
                   f"Vérifiez que les données DPE ont été collectées pour ce département.")
    else:
        # ── KPIs DPE header ──
        stats_dpe = get_dpe_stats(dept)
        kc1, kc2, kc3, kc4, kc5 = st.columns(5)
        kc1.metric("Communes couvertes", f"{stats_dpe.get('nb_communes', 0)}")
        kc2.metric("DPE E/F/G", f"{stats_dpe.get('nb_dpe_total', 0):,}")
        kc3.metric("Passoires F+G", f"{stats_dpe.get('nb_fg', 0):,}")
        kc4.metric("% F+G moyen", f"{stats_dpe.get('pct_fg_moyen', 0):.0f}%")
        nb_tx_dept = len(df)
        kc5.metric("Transactions DVF", f"{nb_tx_dept:,}")

        st.markdown("---")

        # ── Croisement DVF × DPE par commune ──
        # Agréger DVF par commune
        dvf_communes = df.groupby("nom_commune").agg(
            nb_transactions=("valeur_fonciere", "count"),
            prix_m2_median=("prix_m2", "median"),
            prix_m2_moy=("prix_m2", "mean"),
            volume_total=("valeur_fonciere", "sum"),
            surface_moy=("surface_utile", "mean"),
            score_dvf_moy=("score", "mean"),
        ).reset_index()
        dvf_communes.columns = ["commune", "nb_transactions", "prix_m2_median",
                                 "prix_m2_moy", "volume_total", "surface_moy", "score_dvf_moy"]

        # Jointure DVF × DPE
        merged = dvf_communes.merge(
            dpe_communes[["commune", "nb_dpe_efg", "nb_f", "nb_g", "pct_fg", "conso_moy", "ges_moy", "periode_dominante"]],
            on="commune", how="inner"
        )

        if merged.empty:
            st.info("Aucune commune en commun entre DVF et DPE. Vérifiez les noms de communes.")
        else:
            # Score croisé : combine score DVF (opportunité immo) + concentration passoires
            merged["score_dvf"] = merged["score_dvf_moy"].fillna(0).round(0).astype(int)
            _pct_max = merged["pct_fg"].max()
            merged["score_dpe"] = (merged["pct_fg"].fillna(0) * 100 / (_pct_max if _pct_max > 0 else 1)).clip(0, 100).round(0).astype(int)
            merged["score_croise"] = ((merged["score_dvf"] * 0.5 + merged["score_dpe"] * 0.5)
                                      .round(0).astype(int))
            merged = merged.sort_values("score_croise", ascending=False)

            st.markdown(f"**{len(merged)} communes croisées DVF × DPE** — triées par score d'opportunité combiné")

            # ── Top opportunités croisées ──
            col_chart, col_table = st.columns([1, 1])

            with col_chart:
                st.markdown("##### Score croisé par commune")
                import plotly.express as px
                top20 = merged.head(20).copy()
                fig = px.bar(
                    top20, x="score_croise", y="commune",
                    orientation="h",
                    color="pct_fg",
                    color_continuous_scale="YlOrRd",
                    labels={"score_croise": "Score combiné", "commune": "", "pct_fg": "% F+G"},
                    title=f"Top 20 communes — Dept {dept}",
                )
                fig.update_layout(height=500, yaxis=dict(autorange="reversed"))
                st.plotly_chart(fig, use_container_width=True)

            with col_table:
                st.markdown("##### Détail chiffré")
                display_cols = ["commune", "score_croise", "nb_transactions", "prix_m2_median",
                                "nb_dpe_efg", "nb_f", "nb_g", "pct_fg", "conso_moy"]
                st.dataframe(
                    merged[display_cols].head(30),
                    use_container_width=True, height=500,
                    column_config={
                        "score_croise": st.column_config.ProgressColumn("Score", min_value=0, max_value=100, format="%d"),
                        "pct_fg": st.column_config.ProgressColumn("% F+G", min_value=0, max_value=100, format="%.0f%%"),
                    },
                )

            st.markdown("---")

            # ── Scatter plot : prix × passoires ──
            st.markdown("##### Prix vs Passoires thermiques")
            fig2 = px.scatter(
                merged, x="prix_m2_median", y="pct_fg",
                size="nb_transactions", color="score_croise",
                color_continuous_scale="RdYlGn_r",
                hover_name="commune",
                hover_data={"nb_dpe_efg": True, "nb_f": True, "nb_g": True,
                            "conso_moy": True, "score_croise": True},
                labels={"prix_m2_median": "Prix médian €/m²", "pct_fg": "% passoires F+G",
                         "nb_transactions": "Volume transactions", "score_croise": "Score"},
                title="Opportunités : prix bas + forte concentration passoires = haut potentiel rénovation",
            )
            fig2.update_layout(height=450)
            st.plotly_chart(fig2, use_container_width=True)

            st.markdown("---")

            # ── Drill-down commune ──
            st.markdown("##### Détail par commune")
            communes_list = merged["commune"].tolist()
            selected_commune = st.selectbox("Sélectionner une commune", communes_list, key="dpe_commune_select")

            if selected_commune:
                row = merged[merged["commune"] == selected_commune].iloc[0]
                dc1, dc2, dc3, dc4, dc5, dc6 = st.columns(6)
                dc1.metric("Score combiné", f"{row['score_croise']}/100")
                dc2.metric("Transactions DVF", f"{row['nb_transactions']}")
                dc3.metric("Prix médian €/m²", f"{row['prix_m2_median']:,.0f}")
                dc4.metric("DPE E/F/G", f"{row['nb_dpe_efg']}")
                dc5.metric("Passoires F+G", f"{row['nb_f'] + row['nb_g']}")
                dc6.metric("% F+G", f"{row['pct_fg']}%")

                # Charger les logements DPE individuels de cette commune
                with st.expander(f"📍 Logements F+G à {selected_commune}", expanded=False):
                    try:
                        dpe_details = get_dpe_logements(
                            departement=dept,
                            etiquettes=["F", "G"],
                            commune=selected_commune,
                            limit=200
                        )
                        if not dpe_details.empty:
                            st.caption(f"{len(dpe_details)} passoires thermiques trouvées")
                            show_cols = [c for c in ["adresse", "etiquette_dpe", "score_urgence",
                                                      "conso_ef", "emission_ges", "periode_construction",
                                                      "energie_principale", "type_batiment"] if c in dpe_details.columns]
                            st.dataframe(
                                dpe_details[show_cols].sort_values("score_urgence", ascending=False),
                                use_container_width=True, height=350,
                            )

                            # Bouton ajout CRM batch
                            if st.button(f"➕ Ajouter {selected_commune} au CRM", key="dpe_crm_add"):
                                contact = add_contact(
                                    f"Prospect DPE — {selected_commune}",
                                    type_contact="DPE",
                                    notes=f"{row['nb_f']+row['nb_g']} passoires F/G, score {row['score_croise']}"
                                )
                                if contact:
                                    add_opportunite(
                                        contact.get("id", ""),
                                        f"Rénovation {selected_commune} ({row['nb_f']+row['nb_g']} F/G)",
                                        selected_commune, "DPE Zone",
                                        0, 0, row["prix_m2_median"],
                                        row["score_croise"], source="DPE×DVF"
                                    )
                                    st.success(f"✅ {selected_commune} ajouté au CRM")
                        else:
                            st.info("Aucun logement F/G trouvé pour cette commune.")
                    except Exception as e:
                        st.warning(f"Impossible de charger les détails DPE : {e}")

            # ── Export croisement ──
            st.markdown("---")
            csv_cross = merged.to_csv(index=False).encode("utf-8")
            st.download_button(
                "📥 Exporter le croisement DVF × DPE",
                csv_cross,
                f"dvf_dpe_croise_{dept}.csv",
                "text/csv",
                key="dl_dpe_cross"
            )


# ── TAB 6 : CRM PIPELINE (Mobile-first) ──────────────────────────────────────

with tab_crm:

    # ── ACTIONS RAPIDES MOBILE ────────────────────────────────────────────

    def action_buttons(contact: dict, opp_id: str = "") -> None:
        """Boutons clic-to-action pour mobile."""
        tel   = contact.get("tel","").strip()
        email = contact.get("email","").strip()
        nom   = contact.get("nom","")

        btns = []
        if tel:
            btns.append(f'''<a href="tel:{tel}" style="display:inline-flex;align-items:center;gap:.4rem;
            background:#1D9E75;color:#fff;padding:.55rem 1rem;border-radius:6px;text-decoration:none;
            font-size:.85rem;font-weight:500">📞 Appeler</a>''')
            btns.append(f'''<a href="sms:{tel}" style="display:inline-flex;align-items:center;gap:.4rem;
            background:#185FA5;color:#fff;padding:.55rem 1rem;border-radius:6px;text-decoration:none;
            font-size:.85rem;font-weight:500">💬 SMS</a>''')
            wa = tel.replace("+","").replace(" ","")
            btns.append(f'''<a href="https://wa.me/{wa}" target="_blank" style="display:inline-flex;align-items:center;gap:.4rem;
            background:#25D366;color:#fff;padding:.55rem 1rem;border-radius:6px;text-decoration:none;
            font-size:.85rem;font-weight:500">🟢 WhatsApp</a>''')
        if email:
            btns.append(f'''<a href="mailto:{email}?subject=SAHAR — {nom}" style="display:inline-flex;align-items:center;gap:.4rem;
            background:#444;color:#fff;padding:.55rem 1rem;border-radius:6px;text-decoration:none;
            font-size:.85rem;font-weight:500">✉️ Email</a>''')
        if btns:
            st.markdown(
                '<div style="display:flex;gap:.5rem;flex-wrap:wrap;margin:.5rem 0">' + "".join(btns) + '</div>',
                unsafe_allow_html=True
            )

    def dictaphone_widget(opp_id: str, key: str) -> None:
        """Widget dictaphone HTML5 pour compte rendu vocal sur mobile."""
        st.components.v1.html(f"""
        <div style="font-family:sans-serif;padding:.5rem 0">
          <button id="btn_{key}" onclick="toggleRec_{key}()"
            style="background:#E24B4A;color:#fff;border:none;padding:.6rem 1.2rem;
            border-radius:6px;font-size:.9rem;cursor:pointer;font-weight:500">
            🎤 Dicter le compte rendu
          </button>
          <span id="status_{key}" style="font-size:.78rem;color:#888;margin-left:.75rem"></span>
          <div id="result_{key}" style="margin-top:.5rem;padding:.6rem;background:#f5f5f5;
            border-radius:6px;font-size:.88rem;min-height:40px;display:none"></div>
          <button id="copy_{key}" onclick="copyText_{key}()"
            style="display:none;margin-top:.4rem;background:#1a1a1a;color:#fff;border:none;
            padding:.4rem .9rem;border-radius:6px;font-size:.82rem;cursor:pointer">
            📋 Copier
          </button>
        </div>
        <script>
        var rec_{key} = false;
        var recognition_{key} = null;
        var transcript_{key} = "";

        function toggleRec_{key}() {{
          if (!rec_{key}) {{
            if (!window.SpeechRecognition && !window.webkitSpeechRecognition) {{
              document.getElementById('status_{key}').textContent = "Dictaphone non supporté sur ce navigateur";
              return;
            }}
            var SR = window.SpeechRecognition || window.webkitSpeechRecognition;
            recognition_{key} = new SR();
            recognition_{key}.lang = 'fr-FR';
            recognition_{key}.continuous = true;
            recognition_{key}.interimResults = true;
            recognition_{key}.onresult = function(e) {{
              var interim = '';
              var final = '';
              for (var i = e.resultIndex; i < e.results.length; i++) {{
                if (e.results[i].isFinal) final += e.results[i][0].transcript;
                else interim += e.results[i][0].transcript;
              }}
              transcript_{key} += final;
              var div = document.getElementById('result_{key}');
              div.style.display = 'block';
              div.textContent = transcript_{key} + interim;
            }};
            recognition_{key}.onerror = function(e) {{
              document.getElementById('status_{key}').textContent = 'Erreur: ' + e.error;
            }};
            recognition_{key}.start();
            rec_{key} = true;
            document.getElementById('btn_{key}').textContent = '⏹ Arrêter';
            document.getElementById('btn_{key}').style.background = '#888';
            document.getElementById('status_{key}').textContent = '● Enregistrement...';
          }} else {{
            recognition_{key}.stop();
            rec_{key} = false;
            document.getElementById('btn_{key}').textContent = '🎤 Dicter le compte rendu';
            document.getElementById('btn_{key}').style.background = '#E24B4A';
            document.getElementById('status_{key}').textContent = '✓ Terminé';
            document.getElementById('copy_{key}').style.display = 'inline-block';
          }}
        }}

        function copyText_{key}() {{
          navigator.clipboard.writeText(transcript_{key}).then(function() {{
            document.getElementById('status_{key}').textContent = '✓ Copié !';
          }});
        }}
        </script>
        """, height=140)

    # ── ONGLETS CRM ──────────────────────────────────────────────────────

    crm_tabs = st.tabs(["📌 Pipeline", "👥 Contacts", "📅 Activités", "📨 Leads & Séquences", "📊 KPIs"])

    # PIPELINE
    with crm_tabs[0]:
        opps = st.session_state.crm_opportunites
        if not opps:
            st.info("Ajoutez des opportunités depuis l'onglet Transactions.")
        else:
            # Vue mobile : liste par étape avec accordéon
            for stage in STAGES:
                stage_opps = [o for o in opps if o["stage"] == stage]
                valeur = sum(o["prix"] for o in stage_opps)
                nb = len(stage_opps)
                st.markdown(
                    f'''<div style="background:{STAGE_COLORS[stage]};padding:.6rem .9rem;
                    border-radius:6px;margin:.4rem 0;font-size:.88rem">
                    <b>{stage}</b> — {nb} opp.
                    {"  •  " + f"{valeur/1000:.0f}k€" if nb else ""}
                    </div>''', unsafe_allow_html=True
                )
                for opp in stage_opps:
                    sc = opp["score"]
                    sc_col = "#1D9E75" if sc>=70 else "#BA7517" if sc>=40 else "#E24B4A"
                    # Trouver le contact associé
                    contact = next((c for c in st.session_state.crm_contacts
                                    if c["id"] == opp.get("contact_id")), {})

                    with st.expander(f"{opp['titre'][:45]}", expanded=False):
                        # Infos bien
                        st.markdown(f"""
**Adresse :** {opp["adresse"]}
**Type :** {opp["type_bien"]} — {opp["surface"]:.0f} m²
**Prix :** {opp["prix"]:,.0f} € ({opp["prix_m2"]:.0f} €/m²)
**Score :** <span style="color:{sc_col};font-weight:600">{sc}/100</span>
**Contact :** {contact.get("nom","—")} {contact.get("tel","")}
""", unsafe_allow_html=True)

                        # Actions clic-to-action
                        if contact:
                            action_buttons(contact, opp["id"])

                        # Email commercial depuis la fiche
                        with st.expander("✉️ Envoyer un email"):
                            email_type = st.selectbox(
                                "Type d'email",
                                ["Email libre (relance, proposition…)",
                                 "Analyse marché DVF",
                                 "Alerte DPE — logement interdit location"],
                                key=f"etype_{opp['id']}"
                            )
                            if email_type == "Email libre (relance, proposition…)":
                                sujet_e = st.text_input("Objet", key=f"esujet_{opp['id']}")
                                msg_e = st.text_area("Message", height=100, key=f"emsg_{opp['id']}")
                                if st.button("Envoyer", key=f"esend_{opp['id']}", type="primary"):
                                    if contact.get("email") and sujet_e and msg_e:
                                        from shared.automation import email_prospect_generique
                                        ok = email_prospect_generique(
                                            contact["email"], contact["nom"],
                                            sujet_e, msg_e
                                        )
                                        if ok:
                                            crm_db.add_activite(opp["id"], "Email", sujet_e, statut="Fait")
                                            st.success("✅ Email envoyé")
                                            st.rerun()
                                    else:
                                        st.warning("Email, objet et message requis")

                            elif email_type == "Analyse marché DVF":
                                mediane = df[df["nom_commune"]==opp.get("adresse","").split(",")[-1].strip()]["prix_m2"].median() if not df.empty else opp["prix_m2"]
                                st.caption(f"Médiane secteur estimée : {mediane:.0f} €/m²")
                                if st.button("Envoyer analyse DVF", key=f"edvf_{opp['id']}", type="primary"):
                                    if contact.get("email"):
                                        from shared.automation import email_prospect_dvf
                                        ok = email_prospect_dvf(
                                            contact["email"], contact["nom"],
                                            opp["adresse"], opp["adresse"].split(",")[-1].strip(),
                                            opp["prix"], opp["surface"],
                                            opp["prix_m2"], float(mediane), opp["score"]
                                        )
                                        if ok:
                                            crm_db.add_activite(opp["id"], "Email", "Analyse marché DVF envoyée", statut="Fait")
                                            st.success("✅ Email envoyé")
                                            st.rerun()

                            elif email_type == "Alerte DPE — logement interdit location":
                                etiq = st.selectbox("Étiquette DPE", ["G","F","E"], key=f"edpe_etiq_{opp['id']}")
                                if st.button("Envoyer alerte DPE", key=f"edpe_{opp['id']}", type="primary"):
                                    if contact.get("email"):
                                        from shared.automation import email_prospect_dpe
                                        ok = email_prospect_dpe(
                                            contact["email"], contact["nom"],
                                            opp["adresse"], opp["adresse"].split(",")[-1].strip(),
                                            etiq, opp["surface"]
                                        )
                                        if ok:
                                            crm_db.add_activite(opp["id"], "Email", f"Alerte DPE {etiq} envoyée", statut="Fait")
                                            st.success("✅ Email envoyé")
                                            st.rerun()

                        # Changer étape
                        cur_idx = STAGES.index(opp["stage"])
                        new_stage = st.selectbox("Étape", STAGES, index=cur_idx,
                                                  key=f"stg_{opp['id']}")
                        if new_stage != opp["stage"]:
                            crm_db.update_stage(opp["id"], new_stage)
                            st.rerun()

                        # Activité rapide + dictaphone
                        st.markdown("**Ajouter une activité**")
                        col_a1, col_a2 = st.columns([1,2])
                        with col_a1:
                            act_t = st.selectbox("Type",
                                ["Appel","Email","SMS","WhatsApp","Visite","RDV","Note"],
                                key=f"at_{opp['id']}")
                        with col_a2:
                            act_n = st.text_input("Note rapide", key=f"an_{opp['id']}")

                        dictaphone_widget(opp["id"], f"d_{opp['id']}")
                        st.caption("💡 Après dictée : copiez le texte ci-dessus dans la note")

                        if st.button("✓ Enregistrer l'activité", key=f"ab_{opp['id']}", type="primary"):
                            crm_db.add_activite(opp["id"], act_t, act_n)
                            st.success("Activité enregistrée !")
                            st.rerun()

    # CONTACTS
    with crm_tabs[1]:
        # Formulaire compact
        with st.expander("➕ Nouveau contact", expanded=not st.session_state.crm_contacts):
            with st.form("f_contact"):
                nom_f = st.text_input("Nom *")
                c1f, c2f = st.columns(2)
                with c1f: tel_f  = st.text_input("Téléphone")
                with c2f: email_f = st.text_input("Email")
                type_f = st.selectbox("Type", ["Vendeur","Acheteur","Investisseur","Promoteur","Agent","Autre"])
                secteur_f = st.selectbox("Secteur", ["immobilier","energie","retail","autre"])
                notes_f = st.text_area("Notes", height=60)
                send_email_f = st.toggle("Envoyer email de bienvenue", value=True)
                if st.form_submit_button("Créer", type="primary") and nom_f:
                    crm_db.add_contact(nom_f, email_f, tel_f, type_f, notes_f,
                                       envoyer_email_bienvenue=send_email_f,
                                       secteur=secteur_f)
                    st.success(f"✓ {nom_f} créé")
                    st.rerun()

        # Liste contacts avec actions
        for c in st.session_state.crm_contacts:
            nb_opps_c = len([o for o in st.session_state.crm_opportunites if o.get("contact_id") == c["id"]])
            with st.expander(f"**{c['nom']}** — {c['type']} — {nb_opps_c} opp.", expanded=False):
                st.markdown(f"📞 {c.get('tel','—')}  |  ✉️ {c.get('email','—')}")
                if c.get("notes"): st.caption(c["notes"])
                action_buttons(c)

                # Activités liées
                acts_c = [a for a in st.session_state.crm_activites
                           if any(o["id"] == a["opp_id"] and o.get("contact_id") == c["id"]
                                  for o in st.session_state.crm_opportunites)]
                if acts_c:
                    st.caption(f"Dernières activités ({len(acts_c)}) :")
                    for a in sorted(acts_c, key=lambda x: x["date_creation"], reverse=True)[:3]:
                        st.caption(f"  {a['date']} · {a['type']} · {a['notes'][:50]}")

    # ACTIVITÉS
    with crm_tabs[2]:
        acts = st.session_state.crm_activites

        # Formulaire nouvelle activité
        with st.expander("➕ Nouvelle activité"):
            opps_list = [f"{o['id']} — {o['titre'][:40]}" for o in st.session_state.crm_opportunites]
            if opps_list:
                with st.form("f_activite"):
                    opp_s = st.selectbox("Opportunité", opps_list)
                    c1a, c2a = st.columns(2)
                    with c1a:
                        type_a = st.selectbox("Type", ["Appel","Email","SMS","WhatsApp","Visite","RDV","Note","Relance"])
                        stat_a = st.selectbox("Statut", ["À faire","Fait","Annulé"])
                    with c2a:
                        date_a = st.date_input("Date", value=date.today())
                    notes_a = st.text_area("Notes", height=80)
                    if st.form_submit_button("Créer", type="primary"):
                        crm_db.add_activite(opp_s.split(" — ")[0], type_a, notes_a,
                                             date_a.strftime("%d/%m/%Y"), stat_a)
                        st.success("Activité créée !")
                        st.rerun()
            else:
                st.info("Créez d'abord une opportunité.")

        # Liste activités par date
        if acts:
            for a in sorted(acts, key=lambda x: x["date_creation"], reverse=True):
                opp_ref = next((o["titre"][:30] for o in st.session_state.crm_opportunites
                                if o["id"] == a["opp_id"]), a["opp_id"])
                statut_icon = {"Fait":"✅","À faire":"⏳","Annulé":"❌"}.get(a["statut"],"•")
                st.markdown(
                    f'''<div style="border:1px solid #e5e5e5;border-radius:6px;padding:.7rem .9rem;margin:.3rem 0;font-size:.85rem">
                    <b>{statut_icon} {a["type"]}</b> · {a["date"]} · <span style="color:#888">{opp_ref}</span>
                    <div style="color:#444;margin-top:.3rem">{a.get("notes","")}</div>
                    </div>''', unsafe_allow_html=True
                )
        else:
            st.info("Aucune activité.")

    # LEADS & SÉQUENCES + CAMPAGNES
    with crm_tabs[3]:
        st.subheader("Leads, séquences & campagnes")

        # ── SECTION CAMPAGNE BATCH EMAIL ─────────────────────────────────
        with st.expander("📨 Campagne email batch — Prospection automatique", expanded=False):
            st.caption("Envoyez un email commercial à tous les contacts de votre CRM ayant un email.")

            _contacts_with_email = [c for c in st.session_state.crm_contacts if c.get("email")]
            st.markdown(f"**{len(_contacts_with_email)} contacts** avec email dans le CRM")

            if _contacts_with_email:
                _camp_type = st.selectbox("Type de campagne", [
                    "Analyse marché DVF (personnalisée)",
                    "Email libre (même message pour tous)",
                ], key="camp_type")

                if _camp_type == "Analyse marché DVF (personnalisée)":
                    st.caption(
                        "Chaque contact recevra un email avec les données DVF de son opportunité "
                        "(prix, surface, médiane secteur, score). Seuls les contacts liés à au moins une opportunité seront ciblés."
                    )
                    _camp_eligible = []
                    for c in _contacts_with_email:
                        _c_opps = [o for o in st.session_state.crm_opportunites if o.get("contact_id") == c["id"]]
                        if _c_opps:
                            _camp_eligible.append((c, _c_opps[0]))  # première opp
                    st.markdown(f"**{len(_camp_eligible)} contacts éligibles** (avec opportunité + email)")

                    if _camp_eligible and st.button(f"🚀 Lancer la campagne DVF ({len(_camp_eligible)} emails)", type="primary", key="camp_dvf_go"):
                        _sent = 0
                        _errors = 0
                        _progress = st.progress(0)
                        for idx, (contact, opp) in enumerate(_camp_eligible):
                            try:
                                from shared.automation import email_prospect_dvf
                                _commune = opp.get("adresse", "").split(",")[-1].strip() if "," in opp.get("adresse","") else ""
                                _mediane = df[df["nom_commune"] == _commune]["prix_m2"].median() if _commune and not df.empty else opp.get("prix_m2", 0)
                                _mediane = float(_mediane) if pd.notna(_mediane) else float(opp.get("prix_m2", 0))
                                ok = email_prospect_dvf(
                                    contact["email"], contact["nom"],
                                    opp.get("adresse",""), _commune,
                                    opp.get("prix",0), opp.get("surface",0),
                                    opp.get("prix_m2",0), _mediane, opp.get("score",0)
                                )
                                if ok:
                                    crm_db.add_activite(opp["id"], "Email", "Campagne DVF batch", statut="Fait")
                                    _sent += 1
                                else:
                                    _errors += 1
                            except Exception as e:
                                _errors += 1
                            _progress.progress((idx + 1) / len(_camp_eligible))
                        st.success(f"✅ Campagne terminée : {_sent} envoyés, {_errors} erreurs")

                else:  # Email libre batch
                    _camp_sujet = st.text_input("Objet de l'email", key="camp_sujet")
                    _camp_msg = st.text_area("Message (le prénom sera inséré automatiquement)", height=120, key="camp_msg")

                    if _camp_sujet and _camp_msg and st.button(f"🚀 Envoyer à {len(_contacts_with_email)} contacts", type="primary", key="camp_libre_go"):
                        _sent = 0
                        _errors = 0
                        _progress = st.progress(0)
                        for idx, c in enumerate(_contacts_with_email):
                            try:
                                from shared.automation import email_prospect_generique
                                ok = email_prospect_generique(
                                    c["email"], c["nom"],
                                    _camp_sujet, _camp_msg
                                )
                                if ok:
                                    _sent += 1
                                    # Logger l'activité si opportunité liée
                                    _c_opp = next((o for o in st.session_state.crm_opportunites
                                                   if o.get("contact_id") == c["id"]), None)
                                    if _c_opp:
                                        crm_db.add_activite(_c_opp["id"], "Email", f"Campagne : {_camp_sujet[:30]}", statut="Fait")
                                else:
                                    _errors += 1
                            except Exception:
                                _errors += 1
                            _progress.progress((idx + 1) / len(_contacts_with_email))
                        st.success(f"✅ Campagne terminée : {_sent} envoyés, {_errors} erreurs")
            else:
                st.info("Aucun contact avec email. Ajoutez des contacts avec email depuis les transactions.")

        st.markdown("---")

        leads = st.session_state.get("crm_leads", [])

        # Ajouter lead manuellement
        with st.expander("➕ Nouveau lead manuel"):
            with st.form("f_lead"):
                c1l, c2l = st.columns(2)
                with c1l:
                    nom_l = st.text_input("Nom *")
                    email_l = st.text_input("Email *")
                with c2l:
                    tel_l = st.text_input("Téléphone")
                    secteur_l = st.selectbox("Secteur",
                        ["Immobilier","Énergie / Rénovation","Retail / Franchise",
                         "RH / Recrutement","Automobile","Autre"])
                msg_l = st.text_area("Message", height=60)
                if st.form_submit_button("Créer + envoyer J+0", type="primary"):
                    if nom_l and email_l:
                        crm_db.add_lead(nom_l, email_l, tel_l, secteur_l, msg_l)
                        st.success(f"✓ Lead créé — email J+0 envoyé à {email_l}")
                        st.rerun()

        # Liste leads
        if leads:
            for lead in sorted(leads, key=lambda x: x.get("created_at",""), reverse=True)[:20]:
                statut_icon = {"nouveau":"🆕","contacté":"📞","qualifié":"✅","perdu":"❌"}.get(lead.get("statut","nouveau"),"•")
                with st.expander(f"{statut_icon} {lead['nom']} — {lead.get('secteur','')} — {lead.get('email','')}"):
                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown(f"**Email :** {lead.get('email','—')}")
                        st.markdown(f"**Tél :** {lead.get('tel','—')}")
                        st.markdown(f"**Source :** {lead.get('source','—')}")
                    with c2:
                        nouveau_statut = st.selectbox(
                            "Statut", ["nouveau","contacté","qualifié","perdu"],
                            index=["nouveau","contacté","qualifié","perdu"].index(lead.get("statut","nouveau")),
                            key=f"lstatut_{lead['id']}"
                        )
                        if nouveau_statut != lead.get("statut"):
                            crm_db.update_lead_statut(lead["id"], nouveau_statut)
                            st.rerun()

                    # Boutons séquence
                    st.markdown("**Séquence email**")
                    col_s1, col_s2, col_s3, col_s4 = st.columns(4)
                    with col_s1:
                        if st.button("J+0", key=f"j0_{lead['id']}"):
                            from shared.emails_site import envoyer_j0
                            ok = envoyer_j0(lead["nom"], lead["email"], lead.get("secteur",""))
                            st.success("✅ J+0 envoyé") if ok else st.error("Erreur")
                    with col_s2:
                        if st.button("J+3", key=f"j3_{lead['id']}"):
                            from shared.emails_site import envoyer_j3
                            ok = envoyer_j3(lead["nom"], lead["email"], lead.get("secteur",""))
                            st.success("✅ J+3 envoyé") if ok else st.error("Erreur")
                    with col_s3:
                        if st.button("J+7", key=f"j7_{lead['id']}"):
                            from shared.emails_site import envoyer_j7
                            ok = envoyer_j7(lead["nom"], lead["email"], lead.get("secteur",""))
                            st.success("✅ J+7 envoyé") if ok else st.error("Erreur")
                    with col_s4:
                        if st.button("Démo", key=f"demo_{lead['id']}"):
                            st.session_state[f"show_demo_{lead['id']}"] = True

                    if st.session_state.get(f"show_demo_{lead['id']}"):
                        with st.form(f"f_demo_{lead['id']}"):
                            date_d = st.text_input("Date & heure", placeholder="Jeudi 3 avril à 14h30")
                            lien_d = st.text_input("Lien visio (optionnel)")
                            if st.form_submit_button("Confirmer la démo →"):
                                from shared.emails_site import confirmer_demo
                                ok = confirmer_demo(lead["nom"], lead["email"],
                                                    date_d, lien_d, lead.get("secteur",""))
                                if ok:
                                    st.success("✅ Confirmation envoyée")
                                    st.session_state.pop(f"show_demo_{lead['id']}", None)
                                    crm_db.update_lead_statut(lead["id"], "contacté")
                                    st.rerun()
        else:
            st.info("Aucun lead. Ils apparaissent ici quand quelqu'un remplit le formulaire du site.")

    # KPIs
    with crm_tabs[4]:
        opps_all = st.session_state.crm_opportunites
        acts_all  = st.session_state.crm_activites
        contacts_all = st.session_state.crm_contacts
        leads_all = st.session_state.get("crm_leads", [])

        # ROW 1 : Métriques principales
        _k1, _k2, _k3, _k4, _k5, _k6 = st.columns(6)
        with _k1: st.metric("Contacts", len(contacts_all))
        with _k2: st.metric("Opportunités", len(opps_all))
        with _k3: st.metric("Activités", len(acts_all))
        _val = sum(o["prix"] for o in opps_all)
        with _k4: st.metric("Pipeline", f"{_val/1000:.0f}k€" if _val else "0€")
        _closing_n = len([o for o in opps_all if o["stage"]=="Closing"])
        with _k5: st.metric("En closing", _closing_n)
        _sc_moy = sum(o["score"] for o in opps_all)/len(opps_all) if opps_all else 0
        with _k6: st.metric("Score moyen", f"{_sc_moy:.0f}/100")

        # ROW 2 : Conversion & Performance
        st.markdown("---")
        st.markdown("**📈 Indicateurs de conversion**")

        _detecte_n = len([o for o in opps_all if o["stage"]=="Détecté"])
        _contacte_n = len([o for o in opps_all if o["stage"]=="Contacté"])
        _qualifie_n = len([o for o in opps_all if o["stage"]=="Qualifié"])
        _proposition_n = len([o for o in opps_all if o["stage"]=="Proposition"])

        _taux_contact_kpi = round((_contacte_n + _qualifie_n + _proposition_n + _closing_n) / max(len(opps_all), 1) * 100)
        _taux_closing_kpi = round(_closing_n / max(len(opps_all), 1) * 100)
        _acts_par_opp = round(len(acts_all) / max(len(opps_all), 1), 1)
        _val_moy = round(_val / max(len(opps_all), 1) / 1000, 1) if opps_all else 0
        _leads_nouveaux = len([l for l in leads_all if l.get("statut") == "nouveau"])
        _leads_conv = len([l for l in leads_all if l.get("statut") in ["contacté", "qualifié"]])

        _c1, _c2, _c3, _c4, _c5, _c6 = st.columns(6)
        with _c1: st.metric("Taux prise contact", f"{_taux_contact_kpi}%")
        with _c2: st.metric("Taux closing", f"{_taux_closing_kpi}%")
        with _c3: st.metric("Activités/opp", f"{_acts_par_opp}")
        with _c4: st.metric("Valeur moy./opp", f"{_val_moy}k€")
        with _c5: st.metric("Leads à traiter", _leads_nouveaux)
        with _c6: st.metric("Leads convertis", _leads_conv)

        # Répartition pipeline
        if opps_all:
            st.markdown("---")
            st.markdown("**Répartition pipeline**")
            for stage in STAGES:
                nb = len([o for o in opps_all if o["stage"]==stage])
                pct = nb/len(opps_all)*100 if opps_all else 0
                _stage_val = sum(o["prix"] for o in opps_all if o["stage"]==stage)
                st.markdown(
                    f'''<div style="display:flex;align-items:center;gap:.75rem;margin:.3rem 0;font-size:.85rem">
                    <span style="min-width:90px">{stage}</span>
                    <div style="flex:1;height:8px;background:#e5e5e5;border-radius:4px;overflow:hidden">
                      <div style="width:{pct:.0f}%;height:100%;background:#185FA5"></div>
                    </div>
                    <span style="color:#888;min-width:120px;text-align:right">{nb} opp · {_stage_val/1000:.0f}k€</span>
                    </div>''', unsafe_allow_html=True
                )

            # Top 5 opportunités à traiter en priorité
            st.markdown("---")
            st.markdown("**🎯 Priorités d'action**")
            _priority_opps = sorted(
                [o for o in opps_all if o["stage"] in ["Détecté", "Contacté"]],
                key=lambda o: o.get("score", 0), reverse=True
            )[:5]
            if _priority_opps:
                for _po in _priority_opps:
                    _po_acts = len([a for a in acts_all if a.get("opp_id") == _po["id"]])
                    _po_sc_col = "#1D9E75" if _po["score"]>=70 else "#BA7517" if _po["score"]>=40 else "#E24B4A"
                    _po_contact = next((c for c in contacts_all if c["id"] == _po.get("contact_id")), {})
                    st.markdown(
                        f'<div style="border:1px solid #e5e5e5;border-radius:6px;padding:.6rem .9rem;margin:.3rem 0;font-size:.83rem">'
                        f'<b>{_po["titre"][:40]}</b> — '
                        f'<span style="color:{_po_sc_col};font-weight:600">Score {_po["score"]}</span> — '
                        f'{_po["stage"]} — {_po_acts} activités — '
                        f'Contact : {_po_contact.get("nom","—")} {_po_contact.get("tel","")}'
                        f'</div>', unsafe_allow_html=True
                    )
            else:
                st.caption("Toutes les opportunités sont en phase avancée.")


# ── TAB 5 : EXPORT ────────────────────────────────────────────────────────────

with tab_export:
    st.subheader("Export des données")

    col_ex1, col_ex2 = st.columns(2)

    with col_ex1:
        st.markdown("**Export transactions filtrées**")
        st.caption(f"{len(df):,} transactions avec les filtres actuels")

        df_exp = df[[
            'score','date_mutation','nom_commune','adresse','code_postal',
            'type_local','surface_utile','nombre_pieces_principales',
            'surface_terrain','valeur_fonciere','prix_m2','id_parcelle',
            'longitude','latitude'
        ]].copy()
        df_exp['date_mutation'] = df_exp['date_mutation'].dt.strftime('%d/%m/%Y')
        df_exp = df_exp.rename(columns={
            'score':'Score','date_mutation':'Date vente','nom_commune':'Commune',
            'adresse':'Adresse','code_postal':'Code postal','type_local':'Type',
            'surface_utile':'Surface m²','nombre_pieces_principales':'Pièces',
            'surface_terrain':'Terrain m²','valeur_fonciere':'Prix €',
            'prix_m2':'Prix €/m²','id_parcelle':'Réf. parcelle',
            'longitude':'Longitude','latitude':'Latitude'
        })

        csv_data = df_exp.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')
        st.download_button(
            "📥 Télécharger CSV",
            data=csv_data,
            file_name=f"sahar_dvf_{dept}_{datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv",
        )

        try:
            import io, openpyxl
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                df_exp.to_excel(writer, sheet_name='DVF Transactions', index=False)
                ws = writer.sheets['DVF Transactions']
                from openpyxl.styles import Font, PatternFill
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="185FA5", end_color="185FA5", fill_type="solid")
                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width = min(
                        max(len(str(col[0].value or '')), max(len(str(c.value or '')) for c in col)), 40
                    )
            buf.seek(0)
            st.download_button(
                "📥 Télécharger Excel",
                data=buf.getvalue(),
                file_name=f"sahar_dvf_{dept}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except ImportError:
            st.caption("Excel disponible avec : pip install openpyxl")

    with col_ex2:
        st.markdown("**Export CRM**")
        if st.session_state.crm_opportunites:
            df_crm = pd.DataFrame(st.session_state.crm_opportunites)
            csv_crm = df_crm.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')
            st.download_button(
                "📥 Export opportunités CRM (CSV)",
                data=csv_crm,
                file_name=f"sahar_crm_opportunites_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv",
            )
        if st.session_state.crm_contacts:
            df_c = pd.DataFrame(st.session_state.crm_contacts)
            csv_c = df_c.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')
            st.download_button(
                "📥 Export contacts CRM (CSV)",
                data=csv_c,
                file_name=f"sahar_crm_contacts_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv",
            )
        if not st.session_state.crm_opportunites and not st.session_state.crm_contacts:
            st.info("Aucune donnée CRM à exporter.")

        st.markdown("---")
        st.markdown("**Statistiques d'export**")
        st.metric("Transactions disponibles", f"{len(df_raw):,}")
        st.metric("Transactions filtrées", f"{len(df):,}")
        st.metric("Opportunités CRM", len(st.session_state.crm_opportunites))
        st.metric("Contacts CRM", len(st.session_state.crm_contacts))



# ── TAB PILOTAGE ──────────────────────────────────────────────────────────────

with tab_pilotage:
    try:
        import plotly.express as px
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots
    except ImportError:
        st.error("Installer plotly : pip install plotly")
        st.stop()

    opps_p   = st.session_state.crm_opportunites
    acts_p   = st.session_state.crm_activites
    contacts_p = st.session_state.crm_contacts

    # ── HEADER KPIs ──────────────────────────────────────────────────────

    st.markdown("### 🎯 Tableau de bord commercial")

    val_pipeline = sum(o["prix"] for o in opps_p) if opps_p else 0
    val_closing  = sum(o["prix"] for o in opps_p if o["stage"]=="Closing") if opps_p else 0
    nb_acts_total = len(acts_p)
    win_rate = round(len([o for o in opps_p if o["stage"]=="Closing"]) / len(opps_p) * 100) if opps_p else 0

    # KPIs avancés conversion
    _detecte = len([o for o in opps_p if o["stage"]=="Détecté"])
    _contacte = len([o for o in opps_p if o["stage"]=="Contacté"])
    _qualifie = len([o for o in opps_p if o["stage"]=="Qualifié"])
    _proposition = len([o for o in opps_p if o["stage"]=="Proposition"])
    _closing = len([o for o in opps_p if o["stage"]=="Closing"])
    _taux_contact = round(_contacte / max(_detecte + _contacte, 1) * 100) if opps_p else 0
    _taux_qualif = round((_qualifie + _proposition + _closing) / max(len(opps_p), 1) * 100) if opps_p else 0
    _score_moy_pipeline = round(sum(o["score"] for o in opps_p) / len(opps_p)) if opps_p else 0
    _val_moy_opp = round(val_pipeline / max(len(opps_p), 1))

    # ROW 1 : KPIs principaux
    k1,k2,k3,k4,k5 = st.columns(5)
    with k1: st.metric("Contacts", len(contacts_p))
    with k2: st.metric("Opportunités", len(opps_p))
    with k3: st.metric("Pipeline", f"{val_pipeline/1000:.0f}k€")
    with k4: st.metric("En closing", f"{val_closing/1000:.0f}k€")
    with k5: st.metric("Win rate", f"{win_rate}%")

    # ROW 2 : KPIs conversion
    k6,k7,k8,k9,k10 = st.columns(5)
    with k6: st.metric("Taux contact", f"{_taux_contact}%",
                         help="% d'opportunités qui passent de Détecté à Contacté")
    with k7: st.metric("Taux qualification", f"{_taux_qualif}%",
                         help="% d'opps passées au moins en Qualifié")
    with k8: st.metric("Score moyen pipeline", f"{_score_moy_pipeline}/100")
    with k9: st.metric("Valeur moy./opp", f"{_val_moy_opp/1000:.0f}k€")
    with k10: st.metric("Activités totales", nb_acts_total)

    # ── ALERTES AUTOMATIQUES ─────────────────────────────────────────────
    _alertes = []
    # Opportunités stagnantes (Détecté depuis > 7 jours sans activité)
    for o in opps_p:
        if o["stage"] == "Détecté":
            _o_acts = [a for a in acts_p if a.get("opp_id") == o["id"]]
            if not _o_acts:
                _alertes.append(f"⚠️ **{o['titre'][:35]}** — Détecté sans activité, à contacter")
    # Pipeline vide
    if not opps_p:
        _alertes.append("📭 Pipeline vide — ajoutez des opportunités depuis l'onglet Transactions")
    # Pas d'activités récentes
    if nb_acts_total == 0 and opps_p:
        _alertes.append("📞 Aucune activité enregistrée — commencez à contacter vos prospects")
    # Opportunités haute valeur en Détecté
    _high_val_detecte = [o for o in opps_p if o["stage"] == "Détecté" and o.get("score", 0) >= 70]
    if _high_val_detecte:
        _alertes.append(f"🎯 **{len(_high_val_detecte)}** opportunités score ≥ 70 encore en Détecté — action prioritaire")

    if _alertes:
        st.markdown(
            '<div style="background:#fff8e1;border:1px solid #ffe082;border-radius:8px;'
            'padding:.75rem 1rem;margin:.5rem 0">'
            '<div style="font-weight:600;font-size:.85rem;margin-bottom:.4rem">🔔 Alertes commerciales</div>'
            + "<br>".join(f'<span style="font-size:.82rem">{a}</span>' for a in _alertes[:5])
            + '</div>', unsafe_allow_html=True
        )

    st.markdown("---")

    # ── ROW 1 : PIPELINE FUNNEL + ACTIVITÉS ──────────────────────────────

    col_f1, col_f2 = st.columns([1, 1])

    with col_f1:
        st.markdown("**Funnel pipeline**")
        if opps_p:
            stages_counts = {s: len([o for o in opps_p if o["stage"]==s]) for s in STAGES}
            stages_val    = {s: sum(o["prix"] for o in opps_p if o["stage"]==s)/1000 for s in STAGES}

            fig_funnel = go.Figure(go.Funnel(
                y=STAGES,
                x=[stages_counts[s] for s in STAGES],
                texttemplate="%{value} opp.<br>%{percentInitial:.0%}",
                textposition="inside",
                marker=dict(color=["#e5e5e5","#fff3cd","#cfe2ff","#d1ecf1","#d4edda"]),
            ))
            fig_funnel.update_layout(
                height=280, margin=dict(l=0,r=0,t=10,b=0),
                plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                font=dict(size=11)
            )
            st.plotly_chart(fig_funnel, use_container_width=True)

            # Valeur par étape
            for s in STAGES:
                pct = stages_counts[s]/len(opps_p)*100 if opps_p else 0
                st.markdown(
                    f'<div style="display:flex;justify-content:space-between;'
                    f'font-size:.82rem;padding:.2rem 0;border-bottom:1px solid #f0f0f0">'
                    f'<span>{s}</span>'
                    f'<span style="color:#888">{stages_counts[s]} opp · {stages_val[s]:.0f}k€</span>'
                    f'</div>',
                    unsafe_allow_html=True
                )
        else:
            st.info("Aucune opportunité dans le pipeline.")

    with col_f2:
        st.markdown("**Activités par type**")
        if acts_p:
            df_acts_type = pd.DataFrame(acts_p)
            counts = df_acts_type["type"].value_counts().reset_index()
            counts.columns = ["Type","Nb"]
            fig_acts = px.bar(counts, x="Nb", y="Type", orientation="h",
                              color="Nb", color_continuous_scale="Blues",
                              labels={"Nb":"","Type":""})
            fig_acts.update_layout(
                height=200, margin=dict(l=0,r=0,t=10,b=0),
                plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                coloraxis_showscale=False, font=dict(size=11)
            )
            st.plotly_chart(fig_acts, use_container_width=True)

            st.markdown("**Statut activités**")
            for statut in ["À faire","Fait","Annulé"]:
                nb = len([a for a in acts_p if a.get("statut")==statut])
                icon = {"À faire":"⏳","Fait":"✅","Annulé":"❌"}.get(statut,"•")
                st.markdown(
                    f'<div style="display:flex;justify-content:space-between;font-size:.83rem;padding:.2rem 0">'
                    f'<span>{icon} {statut}</span><span style="color:#888">{nb}</span></div>',
                    unsafe_allow_html=True
                )
        else:
            st.info("Aucune activité enregistrée.")

    st.markdown("---")

    # ── ROW 2 : MARCHÉ IMMOBILIER (données DVF) ───────────────────────────

    st.markdown("**Analyse marché — données DVF**")

    col_m1, col_m2, col_m3 = st.columns(3)

    with col_m1:
        # Prix médian par commune top 10
        top_comm = (df.groupby("nom_commune")["prix_m2"]
                    .agg(["median","count"])
                    .query("count >= 10")
                    .nlargest(10,"median")
                    .reset_index())
        top_comm.columns = ["Commune","Médiane €/m²","Nb ventes"]
        fig_comm = px.bar(top_comm, x="Médiane €/m²", y="Commune", orientation="h",
                          title="Top 10 communes — prix médian",
                          color="Médiane €/m²", color_continuous_scale="Blues")
        fig_comm.update_layout(
            height=280, margin=dict(l=0,r=0,t=30,b=0),
            plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
            coloraxis_showscale=False, font=dict(size=11), showlegend=False
        )
        st.plotly_chart(fig_comm, use_container_width=True)

    with col_m2:
        # Évolution mensuelle prix médian
        df_trend = df.groupby("mois")["prix_m2"].median().reset_index()
        df_trend.columns = ["Mois","Prix médian €/m²"]
        fig_trend = px.area(df_trend, x="Mois", y="Prix médian €/m²",
                             title="Tendance prix marché",
                             color_discrete_sequence=["#185FA5"])
        fig_trend.update_layout(
            height=280, margin=dict(l=0,r=0,t=30,b=0),
            plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
            font=dict(size=11)
        )
        st.plotly_chart(fig_trend, use_container_width=True)

    with col_m3:
        # Répartition Appart/Maison + score
        df_type = df.groupby("type_local").agg(
            nb=("prix_m2","count"),
            prix_med=("prix_m2","median"),
            score_med=("score","mean")
        ).reset_index()

        fig_type = go.Figure()
        fig_type.add_trace(go.Bar(
            name="Volume",
            x=df_type["type_local"], y=df_type["nb"],
            marker_color=["#185FA5","#1D9E75"],
            yaxis="y"
        ))
        fig_type.add_trace(go.Scatter(
            name="Prix médian €/m²",
            x=df_type["type_local"], y=df_type["prix_med"],
            mode="markers+lines",
            marker=dict(size=10, color="#E24B4A"),
            yaxis="y2"
        ))
        fig_type.update_layout(
            title="Volume & prix par type",
            height=280, margin=dict(l=0,r=0,t=30,b=0),
            plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
            font=dict(size=11),
            yaxis=dict(title="Nb transactions"),
            yaxis2=dict(title="€/m²", overlaying="y", side="right"),
            legend=dict(orientation="h", y=-0.15)
        )
        st.plotly_chart(fig_type, use_container_width=True)

    st.markdown("---")

    # ── ROW 3 : OPPORTUNITÉS + SCORING ────────────────────────────────────

    col_o1, col_o2 = st.columns([1,1])

    with col_o1:
        st.markdown("**Opportunités à traiter — score ≥ 70**")
        top_opps = df[df["score"] >= 70].nlargest(10, "score")[[
            "score","nom_commune","adresse","type_local",
            "surface_utile","prix_m2","valeur_fonciere"
        ]].copy()
        if not top_opps.empty:
            top_opps = top_opps.rename(columns={
                "score":"Score","nom_commune":"Commune","adresse":"Adresse",
                "type_local":"Type","surface_utile":"m²",
                "prix_m2":"€/m²","valeur_fonciere":"Prix"
            })
            st.dataframe(
                top_opps, use_container_width=True, hide_index=True,
                column_config={
                    "Score": st.column_config.ProgressColumn("Score",min_value=0,max_value=100,format="%d"),
                    "Prix": st.column_config.NumberColumn("Prix",format="%d €"),
                    "€/m²": st.column_config.NumberColumn("€/m²",format="%d €"),
                }
            )
        else:
            st.info("Aucune opportunité score ≥ 70 avec les filtres actuels.")

    with col_o2:
        st.markdown("**Distribution des scores**")
        fig_score = px.histogram(
            df, x="score", nbins=20, color="type_local",
            color_discrete_map={"Appartement":"#185FA5","Maison":"#1D9E75"},
            labels={"score":"Score","count":"Nb","type_local":"Type"},
            barmode="overlay", opacity=0.75
        )
        fig_score.add_vline(x=70, line_dash="dash", line_color="#1D9E75",
                            annotation_text="Seuil fort")
        fig_score.add_vline(x=40, line_dash="dash", line_color="#BA7517",
                            annotation_text="Seuil moyen")
        fig_score.update_layout(
            height=250, margin=dict(l=0,r=0,t=10,b=0),
            plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
            font=dict(size=11), legend=dict(orientation="h",y=-0.2)
        )
        st.plotly_chart(fig_score, use_container_width=True)

        # Stats résumées
        st.markdown("**Résumé marché filtré**")
        c1r, c2r = st.columns(2)
        with c1r:
            st.metric("Prix min €/m²", f"{df['prix_m2'].min():.0f}")
            st.metric("Prix max €/m²", f"{df['prix_m2'].max():.0f}")
        with c2r:
            st.metric("Surface min", f"{df['surface_utile'].min():.0f} m²")
            st.metric("Surface max", f"{df['surface_utile'].max():.0f} m²")

    st.markdown("---")

    # ── ROW 4 : TIMELINE ACTIVITÉS CRM ────────────────────────────────────

    if acts_p:
        st.markdown("**Timeline activités**")
        df_timeline = pd.DataFrame(acts_p)
        if "date" in df_timeline.columns:
            df_timeline["date_dt"] = pd.to_datetime(df_timeline["date"], format="%d/%m/%Y", errors="coerce")
            df_day = (df_timeline.dropna(subset=["date_dt"])
                      .groupby([df_timeline["date_dt"].dt.date,"type"])
                      .size().reset_index(name="nb"))
            df_day.columns = ["Date","Type","Nb"]
            if not df_day.empty:
                fig_tl = px.bar(df_day, x="Date", y="Nb", color="Type",
                                labels={"Nb":"Activités","Date":""},
                                color_discrete_sequence=px.colors.qualitative.Set2)
                fig_tl.update_layout(
                    height=200, margin=dict(l=0,r=0,t=10,b=0),
                    plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                    font=dict(size=11), legend=dict(orientation="h",y=-0.25)
                )
                st.plotly_chart(fig_tl, use_container_width=True)

# ── FOOTER ────────────────────────────────────────────────────────────────────

st.markdown("---")
st.caption(
    f"SAHAR Conseil — Sources : DVF data.gouv.fr, INSEE. "
    f"Données à titre indicatif. Dernière mise à jour filtres : {datetime.now().strftime('%d/%m/%Y %H:%M')}"
)

# ─────────────────────────────────────────────────────────────────────────────
# MODULE PILOTAGE — injecté après tab_export
# ─────────────────────────────────────────────────────────────────────────────
