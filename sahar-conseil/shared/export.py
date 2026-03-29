"""
SAHAR Conseil — export.py
Génération d'exports Excel et PDF depuis les DataFrames Streamlit.
"""

import io
import pandas as pd
from datetime import datetime


# ─────────────────────────────────────────────
# EXPORT EXCEL
# ─────────────────────────────────────────────

def export_excel(df: pd.DataFrame, nom_feuille: str = "Données SAHAR") -> bytes:
    """
    Génère un fichier Excel formaté depuis un DataFrame.
    Compatible avec st.download_button.

    Args:
        df: DataFrame à exporter
        nom_feuille: Nom de l'onglet Excel

    Returns:
        bytes du fichier .xlsx

    Utilisation dans Streamlit :
        xlsx_bytes = export_excel(df)
        st.download_button(
            "📥 Exporter Excel",
            data=xlsx_bytes,
            file_name="sahar_export.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=nom_feuille, index=False)

        # Récupérer la feuille pour la formater
        ws = writer.sheets[nom_feuille]

        # En-têtes en gras
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="185FA5", end_color="185FA5", fill_type="solid")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Ajustement automatique des largeurs de colonnes
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

    output.seek(0)
    return output.getvalue()


def export_excel_multi_feuilles(feuilles: dict) -> bytes:
    """
    Génère un Excel avec plusieurs feuilles.

    Args:
        feuilles: dict {nom_feuille: dataframe}

    Returns:
        bytes du fichier .xlsx

    Exemple:
        xlsx = export_excel_multi_feuilles({
            "Toutes transactions": df_complet,
            "Top opportunités": df_top,
            "Résumé par commune": df_resume,
        })
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for nom, df in feuilles.items():
            df.to_excel(writer, sheet_name=nom[:31], index=False)

    output.seek(0)
    return output.getvalue()


# ─────────────────────────────────────────────
# EXPORT PDF (via reportlab)
# ─────────────────────────────────────────────

def export_pdf_rapport(
    titre: str,
    secteur: str,
    commune: str,
    kpis: dict,
    df_top: pd.DataFrame,
    nb_transactions: int = 0,
) -> bytes:
    """
    Génère un rapport PDF professionnel SAHAR.
    Nécessite : pip install reportlab

    Args:
        titre: Titre du rapport (ex: "Analyse marché immobilier")
        secteur: Secteur (ex: "Immobilier")
        commune: Commune analysée
        kpis: dict de KPIs {label: valeur} à afficher en haut du rapport
        df_top: DataFrame des meilleures opportunités
        nb_transactions: Nombre total de transactions analysées

    Returns:
        bytes du fichier .pdf

    Utilisation Streamlit :
        pdf_bytes = export_pdf_rapport(...)
        st.download_button("📄 Télécharger PDF", pdf_bytes, "rapport_sahar.pdf", "application/pdf")
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        raise ImportError(
            "reportlab n'est pas installé. Lancez : pip install reportlab"
        )

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    couleur_primaire = colors.HexColor("#185FA5")
    couleur_accent = colors.HexColor("#1D9E75")

    style_titre = ParagraphStyle(
        "titre",
        parent=styles["Heading1"],
        textColor=couleur_primaire,
        fontSize=20,
        spaceAfter=6,
    )
    style_sous_titre = ParagraphStyle(
        "sous_titre",
        parent=styles["Normal"],
        textColor=colors.HexColor("#73726c"),
        fontSize=11,
        spaceAfter=20,
    )
    style_section = ParagraphStyle(
        "section",
        parent=styles["Heading2"],
        textColor=couleur_primaire,
        fontSize=13,
        spaceBefore=16,
        spaceAfter=8,
    )

    contenu = []

    # En-tête
    contenu.append(Paragraph("SAHAR Conseil", style_sous_titre))
    contenu.append(Paragraph(titre, style_titre))
    contenu.append(Paragraph(
        f"Secteur : {secteur} | Zone : {commune} | "
        f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
        style_sous_titre
    ))
    contenu.append(HRFlowable(width="100%", thickness=2, color=couleur_primaire))
    contenu.append(Spacer(1, 0.5 * cm))

    # KPIs
    contenu.append(Paragraph("Indicateurs clés", style_section))
    kpi_data = [["Indicateur", "Valeur"]] + [[k, str(v)] for k, v in kpis.items()]
    kpi_table = Table(kpi_data, colWidths=[10 * cm, 6 * cm])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), couleur_primaire),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1EFE8")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D3D1C7")),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    contenu.append(kpi_table)
    contenu.append(Spacer(1, 0.5 * cm))

    # Top opportunités
    if not df_top.empty:
        contenu.append(Paragraph("Top opportunités détectées", style_section))

        # Limiter à 15 colonnes max
        df_export = df_top.head(15).copy()
        header = list(df_export.columns)
        rows = [header] + df_export.astype(str).values.tolist()

        col_width = 17 * cm / max(len(header), 1)
        opp_table = Table(rows, colWidths=[col_width] * len(header), repeatRows=1)
        opp_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), couleur_accent),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EAF3DE")]),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D3D1C7")),
            ("PADDING", (0, 0), (-1, -1), 4),
        ]))
        contenu.append(opp_table)

    # Pied de page
    contenu.append(Spacer(1, 1 * cm))
    contenu.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#D3D1C7")))
    contenu.append(Paragraph(
        f"Rapport généré par SAHAR Conseil — {nb_transactions:,} transactions analysées. "
        "Sources : data.gouv.fr, INSEE, ADEME.",
        ParagraphStyle("footer", parent=styles["Normal"],
                       textColor=colors.HexColor("#888780"), fontSize=8)
    ))

    doc.build(contenu)
    output.seek(0)
    return output.getvalue()
